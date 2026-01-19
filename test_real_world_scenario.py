#!/usr/bin/env python3
"""
Real-World Scenario Test: Simulating actual user workflow

This test simulates the exact scenario the user described:
1. User updates sun simulator master list externally (in Excel)
2. Application should detect the change in real-time
3. New serials should be immediately available for validation
"""

import time
from pathlib import Path
from openpyxl import Workbook, load_workbook
from app.path_utils import get_base_dir
from app.serial_database import SerialDatabase

def simulate_user_workflow():
    """Simulate the exact user workflow"""
    print("="*70)
    print("🔬 REAL-WORLD SCENARIO TEST")
    print("="*70)
    print("\n📝 Scenario: User updates master list externally in Excel")
    print("   Expected: App detects changes in real-time\n")
    
    base_dir = get_base_dir()
    db_file = base_dir / "data" / "PALLETS" / "serial_database.xlsx"
    imported_dir = base_dir / "data" / "IMPORTED DATA"
    master_file = imported_dir / "sun_simulator_data.xlsx"
    
    # Ensure directories exist
    db_file.parent.mkdir(parents=True, exist_ok=True)
    imported_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"📁 Database file: {db_file}")
    print(f"📁 Master file: {master_file}\n")
    
    # STEP 1: Initialize the application (simulating app startup)
    print("="*70)
    print("STEP 1: Application Startup")
    print("="*70)
    print("   Initializing SerialDatabase...")
    serial_db = SerialDatabase(db_file, imported_dir, master_file, defer_init=False)
    print("   ✅ Application initialized\n")
    
    # STEP 2: Create initial database file (simulating existing data)
    print("="*70)
    print("STEP 2: Initial Database State")
    print("="*70)
    print("   Creating database file with existing serials...")
    
    # Create database file (this is what validate_serial checks)
    wb = Workbook()
    ws = wb.active
    ws.title = "SerialNos"
    
    # Headers
    headers = ["SerialNo", "Pm", "Isc", "Voc", "Ipm", "Vpm", "Date", "TTime", "First Imported", "Last Updated"]
    for col, header in enumerate(headers, start=1):
        ws.cell(1, col, header)
    
    # Add initial serials
    initial_serials = ["RW001", "RW002", "RW003"]
    for i, serial in enumerate(initial_serials, start=2):
        ws.cell(i, 1, serial)
        ws.cell(i, 2, 200.0 + (i * 5))  # Varying power
        ws.cell(i, 3, 10.0)
        ws.cell(i, 4, 40.0)
    
    wb.save(db_file)
    wb.close()
    print(f"   ✅ Database file created with {len(initial_serials)} serials")
    print(f"   Serials: {', '.join(initial_serials)}\n")
    
    # STEP 3: Validate existing serials (app is running)
    print("="*70)
    print("STEP 3: App Running - Validate Existing Serials")
    print("="*70)
    for serial in initial_serials:
        result = serial_db.validate_serial(serial)
        status = "✅ FOUND" if result else "❌ NOT FOUND"
        print(f"   Serial '{serial}': {status}")
    print()
    
    # STEP 4: User updates database file externally (simulating Excel edit)
    print("="*70)
    print("STEP 4: USER ACTION - External File Update")
    print("="*70)
    print("   👤 User opens database file in Excel...")
    print("   👤 User adds new serials manually...")
    time.sleep(0.2)  # Simulate user taking time to edit
    
    # Simulate external update to database file
    wb = load_workbook(db_file)
    ws = wb['SerialNos']
    
    # Add new serials (user added these)
    new_serials = ["RW004", "RW005", "RW006"]
    start_row = len(initial_serials) + 2
    
    for i, serial in enumerate(new_serials):
        row = start_row + i
        ws.cell(row, 1, serial)
        ws.cell(row, 2, 220.0 + (i * 5))
        ws.cell(row, 3, 11.0)
        ws.cell(row, 4, 42.0)
    
    wb.save(db_file)  # User saves file
    wb.close()
    print(f"   ✅ User saved database file with {len(new_serials)} new serials")
    print(f"   New serials: {', '.join(new_serials)}\n")
    
    # STEP 5: File change ready for detection
    print("="*70)
    print("STEP 5: File Change Ready for Detection")
    print("="*70)
    time.sleep(0.1)  # Small delay
    
    # Note: We don't manually check has_changed() here because that would
    # consume the change detection. Instead, the next validation will
    # automatically detect the change and refresh the cache.
    print("   ✅ File has been updated externally")
    print("   ✅ Change will be detected automatically on next validation\n")
    
    # STEP 6: Validate new serials (should trigger cache refresh)
    print("="*70)
    print("STEP 6: Validate New Serials (Cache Auto-Refresh)")
    print("="*70)
    print("   App validates new serials (cache automatically refreshes)...")
    print("   (First validation will detect file change and refresh cache)\n")
    
    # First validation will trigger cache refresh
    time.sleep(0.1)  # Small delay to ensure file is fully saved
    
    # Check change count before validation
    change_count_before = serial_db.file_monitor.get_info()['change_count']
    
    all_found = True
    for serial in new_serials:
        result = serial_db.validate_serial(serial)
        status = "✅ FOUND" if result else "❌ NOT FOUND"
        print(f"   Serial '{serial}': {status}")
        if not result:
            all_found = False
    
    # Verify change was detected
    change_count_after = serial_db.file_monitor.get_info()['change_count']
    change_detected = change_count_after > change_count_before
    
    if all_found:
        print(f"\n   ✅ All new serials are immediately available!")
        if change_detected:
            print(f"   ✅ File change was detected (change count: {change_count_before} → {change_count_after})")
        print("   ✅ Cache was automatically refreshed due to file change\n")
    else:
        # Debug: Check what's actually in the file
        print("\n   ⚠️  Some serials not found - checking file contents...")
        try:
            wb = load_workbook(db_file, read_only=True)
            ws = wb['SerialNos']
            file_serials = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    file_serials.append(str(row[0]))
            wb.close()
            print(f"   Serials in file: {file_serials}")
            print(f"   Cache size: {len(serial_db._serial_cache) if serial_db._serial_cache else 0}")
        except Exception as e:
            print(f"   Error checking file: {e}")
        print("   ❌ Cache may not have refreshed properly\n")
        return False
    
    # STEP 7: Verify existing serials still work
    print("="*70)
    print("STEP 7: Verify Existing Serials Still Work")
    print("="*70)
    all_still_work = True
    for serial in initial_serials:
        result = serial_db.validate_serial(serial)
        status = "✅ FOUND" if result else "❌ NOT FOUND"
        print(f"   Serial '{serial}': {status}")
        if not result:
            all_still_work = False
    
    if all_still_work:
        print("\n   ✅ All existing serials still work correctly\n")
    else:
        print("\n   ❌ Some existing serials broken\n")
        return False
    
    # STEP 8: Performance verification
    print("="*70)
    print("STEP 8: Performance Check")
    print("="*70)
    iterations = 100
    start_time = time.time()
    
    for i in range(iterations):
        serial_db.validate_serial(f"PERF{i}")
    
    elapsed = time.time() - start_time
    avg_time = (elapsed / iterations) * 1000
    
    print(f"   {iterations} validations in {elapsed:.4f}s")
    print(f"   Average: {avg_time:.4f}ms per validation")
    print(f"   Throughput: {iterations / elapsed:.0f} validations/second")
    
    if avg_time < 10:
        print("   ✅ Performance: EXCELLENT (no noticeable impact)\n")
    else:
        print("   ⚠️  Performance: May be slower than expected\n")
    
    # Final summary
    print("="*70)
    print("🎉 REAL-WORLD SCENARIO TEST: PASSED")
    print("="*70)
    print("\n📊 Test Results:")
    print("   ✅ File changes detected in real-time")
    print("   ✅ Cache automatically invalidated on file change")
    print("   ✅ New serials immediately available after external update")
    print("   ✅ Existing serials continue to work")
    print(f"   ✅ Performance: {avg_time:.4f}ms per validation (excellent)")
    print("\n💡 Conclusion:")
    print("   The file monitoring system successfully detects external file")
    print("   changes and makes new data available immediately without")
    print("   requiring application restart or manual cache refresh.\n")
    
    return True

if __name__ == "__main__":
    try:
        success = simulate_user_workflow()
        if success:
            print("✅ All tests passed - System is production-ready!")
            exit(0)
        else:
            print("❌ Some tests failed - Review output above")
            exit(1)
    except Exception as e:
        print(f"\n❌ Test failed with error: {e}")
        import traceback
        traceback.print_exc()
        exit(1)

