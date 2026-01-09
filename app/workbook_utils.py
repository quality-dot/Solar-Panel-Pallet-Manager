#!/usr/bin/env python3
"""
Workbook Utilities - Workbook detection and SerialNo validation

Provides functions for finding pallet workbooks and validating SerialNos
against the DATA sheet. Optimized for fast lookups during real-time scanning.
"""

import re
from pathlib import Path
from typing import Optional
from openpyxl import load_workbook

# Import normalize_serial from serial_database for consistent normalization
try:
    from app.serial_database import normalize_serial
except ImportError:
    # Fallback if import fails
    def normalize_serial(serial) -> str:
        """Fallback normalization if import fails"""
        if serial is None:
            return ""
        serial_str = str(serial).strip()
        if not serial_str:
            return ""
        # Remove cell references in parentheses
        serial_str = re.sub(r'\([a-z]\d+\)$', '', serial_str, flags=re.IGNORECASE)
        # Remove .0 suffix from numeric serials
        if serial_str.endswith('.0') and len(serial_str) > 2:
            try:
                float_val = float(serial_str)
                if float_val == int(float_val):
                    serial_str = str(int(float_val))
            except (ValueError, OverflowError):
                pass
        return serial_str.strip()


def find_pallet_workbook(excel_dir: Path, current_workbook_path: Optional[Path] = None) -> Optional[Path]:
    """
    Find the target pallet workbook using existing logic.
    
    Priority:
    1. current_workbook_path (EXCEL/CURRENT.xlsx) if provided and exists
    2. Most recently modified BUILD YYYY Q-X.xlsx file in excel_dir
    
    Args:
        excel_dir: Path to EXCEL directory
        current_workbook_path: Optional path to CURRENT.xlsx file
        
    Returns:
        Path to workbook if found, None otherwise
    """
    # Check for CURRENT.xlsx first
    if current_workbook_path and current_workbook_path.exists():
        return current_workbook_path
    
    # Find BUILD files
    build_files = []
    if not excel_dir.exists():
        return None
    
    for file in excel_dir.glob("*.xlsx"):
        # Skip temp files
        if file.name.startswith("~$"):
            continue
        
        # Check for BUILD pattern
        name = file.name
        if "BUILD" in name.upper():
            # Check for year (4 digits)
            if re.search(r'\d{4}', name):
                # Check for quarter pattern
                if re.search(r'Q\s*-?\s*\d', name, re.IGNORECASE):
                    build_files.append(file)
    
    if not build_files:
        return None
    
    # Return most recently modified
    latest = max(build_files, key=lambda p: p.stat().st_mtime)
    return latest


def validate_serial(serial_no: str, workbook_path: Path) -> bool:
    """
    Validate that a SerialNo exists in the workbook's DATA sheet.
    
    Opens workbook in read-only mode for fast, non-destructive lookup.
    Searches column B (SerialNo column) for the given serial.
    
    Note: DATA sheet is optional. If the workbook doesn't have a DATA sheet,
    this function returns False (serial not found) instead of raising an error.
    This allows the app to work with workbooks that only have a PALLET SHEET.
    
    Args:
        serial_no: Serial number string to validate
        workbook_path: Path to Excel workbook
        
    Returns:
        True if SerialNo found in DATA sheet column B, False otherwise
        (also returns False if DATA sheet doesn't exist)
        
    Raises:
        FileNotFoundError: If workbook doesn't exist
        PermissionError: If workbook is locked/open
    """
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    
    # Open workbook in read-only mode for performance
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except PermissionError:
        # Workbook might be open/locked
        raise PermissionError(f"Workbook is locked or open: {workbook_path}")
    
    try:
        # Get DATA sheet (optional - workbook may only have PALLET SHEET)
        if 'DATA' not in wb.sheetnames:
            # DATA sheet is optional - return False (serial not found) instead of raising error
            # This allows the app to work with workbooks that only have PALLET SHEET
            wb.close()
            return False
        
        data_sheet = wb['DATA']
        
        # Normalize the input serial (handles cell references, .0 suffix, whitespace)
        serial_normalized = normalize_serial(serial_no)
        if not serial_normalized:
            wb.close()
            return False
        
        # Search column B (index 2, 1-based) starting from row 2 (skip header)
        # Column B is index 2 in openpyxl (A=1, B=2)
        for row in data_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            # row is a tuple, first element is column A, second is column B
            if len(row) >= 2 and row[1] is not None:
                # Normalize Excel cell value (handles cell references, .0 suffix, whitespace)
                cell_value = normalize_serial(row[1])
                if cell_value and cell_value == serial_normalized:
                    wb.close()
                    return True
        
        wb.close()
        return False
        
    except Exception as e:
        wb.close()
        raise

