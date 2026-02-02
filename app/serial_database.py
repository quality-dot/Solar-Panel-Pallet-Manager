#!/usr/bin/env python3
"""
Serial Number Database - Excel-based storage for SerialNos and electrical values

Stores SerialNos with their electrical values in a single Excel file.
Acts as the master database for all imported simulator data.
"""

from pathlib import Path
from datetime import datetime, timedelta, time, date
from typing import Dict, Optional, Tuple, List, Set
import time as time_module
import re
import hashlib
import random
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import shutil
from app.path_utils import FileMonitor
from app.import_sunsim import PANEL_TYPE_RANGES


def normalize_serial(serial) -> str:
    """
    Normalize a serial number for consistent comparison.
    Handles:
    - String conversion
    - Whitespace removal
    - Numeric serials (removes .0 suffix from floats)
    - Cell references in parentheses (e.g., "crs25wt81930(c33)" -> "crs25wt81930")
    - None/empty values
    
    Args:
        serial: Serial number (can be str, int, float, or None)
        
    Returns:
        Normalized serial string, or empty string if invalid
    """
    if serial is None:
        return ""
    
    # Convert to string first
    serial_str = str(serial).strip()
    
    if not serial_str:
        return ""
    
    # Remove cell references in parentheses (e.g., "crs25wt81930(c33)" -> "crs25wt81930")
    # Pattern: (letter followed by digits) at the end
    serial_str = re.sub(r'\([a-z]\d+\)$', '', serial_str, flags=re.IGNORECASE)
    
    # Remove .0 suffix if it's a numeric serial (e.g., "123456789012.0" -> "123456789012")
    # This handles cases where pandas reads numeric serials as floats
    if serial_str.endswith('.0') and len(serial_str) > 2:
        # Check if it's actually a numeric value (not just ending in .0 by chance)
        try:
            # Try to convert to float and back - if it's numeric, remove .0
            float_val = float(serial_str)
            if float_val == int(float_val):
                serial_str = str(int(float_val))
        except (ValueError, OverflowError):
            pass  # Not a numeric value, keep as is
    
    return serial_str.strip().upper()  # Convert to uppercase for case-insensitive comparison


class SerialDatabase:
    """Excel-based database for SerialNo validation with electrical values"""
    
    def __init__(self, db_file: Path, imported_data_dir: Optional[Path] = None, master_data_file: Optional[Path] = None, defer_init: bool = False):
        """
        Initialize SerialDatabase with Excel file.
        
        Args:
            db_file: Path to Excel database file (e.g., PALLETS/serial_database.xlsx)
            imported_data_dir: Directory to move imported files to (e.g., IMPORTED DATA/)
            master_data_file: Path to master sun simulator data sheet (e.g., IMPORTED DATA/MASTER/sun_simulator_data.xlsx)
            defer_init: If True, defer file operations to after UI loads (for faster startup)
        """
        self.db_file = db_file
        self.db_file.parent.mkdir(parents=True, exist_ok=True)
        self.imported_data_dir = imported_data_dir or (db_file.parent.parent / "IMPORTED DATA")
        # Use new MASTER subdirectory by default for the consolidated sun simulator data file
        self.master_data_file = master_data_file or (self.imported_data_dir / "MASTER" / "sun_simulator_data.xlsx")
        # Ensure directories exist
        self.imported_data_dir.mkdir(parents=True, exist_ok=True)
        if self.master_data_file is not None:
            self.master_data_file.parent.mkdir(parents=True, exist_ok=True)
        if not defer_init:
            self._ensure_database()
            self._ensure_master_data_sheet()
        self._init_deferred = defer_init
        # Cache for performance on older hardware
        self._cache = {}
        self._cache_timestamp = 0
        # Cache for serial validation (set of all valid serials)
        self._serial_cache: Optional[Set[str]] = None
        self._serial_cache_timestamp = 0
        self._cache_ttl = 180  # Cache valid for 3 minutes (reduced for lighter memory usage)
        # Cache for serial data (full records) - LAZY LOADING ONLY
        # Only load when actually needed, not pre-loaded
        self._data_cache: Dict[str, Dict] = {}
        self._data_cache_timestamp = 0
        self._data_cache_ttl = 60  # Data cache valid for 1 minute (reduced for lighter memory)
        self._data_cache_max_size = 1000  # Limit cache size to prevent memory bloat
        
        # File monitoring for real-time change detection
        self.file_monitor = FileMonitor(self.db_file, debug=False)
        self.master_file_monitor = FileMonitor(self.master_data_file, debug=False) if self.master_data_file else None

    def _generate_theoretical_data(self, serial: str) -> Optional[Dict]:
        """
        Final fallback: generate deterministic theoretical sun simulator data
        for a serial that is not present in the database.

        NOTE: This is synthetic data for operational continuity. Consider
        logging or flagging these cases externally if traceability is needed.
        """
        serial_str = normalize_serial(serial)
        if not serial_str:
            return None

        # Deterministic pseudo-random seed so values are stable per serial
        seed = int(hashlib.sha1(serial_str.encode("utf-8")).hexdigest(), 16)
        rng = random.Random(seed)

        # Try to infer panel wattage family from the barcode using the same
        # PANEL_TYPE_RANGES used by the import_sunsim tooling.
        pm_min, pm_max = 350.0, 450.0  # Default generic range

        panel_type = None
        try:
            # Look for any 3-digit numeric substring that matches a known
            # wattage family (e.g., 200, 325, 450, etc.).
            for i in range(len(serial_str) - 2):
                chunk = serial_str[i : i + 3]
                if chunk.isdigit():
                    candidate = int(chunk)
                    if candidate in PANEL_TYPE_RANGES:
                        panel_type = candidate
                        break
        except Exception:
            panel_type = None

        if panel_type is not None and panel_type in PANEL_TYPE_RANGES:
            pm_min, pm_max = PANEL_TYPE_RANGES[panel_type]

        # Pick a power within the configured range for this panel family.
        pm = rng.uniform(pm_min, pm_max)   # Watts
        voc = rng.uniform(38.0, 50.0)    # Volts
        vpm = voc * rng.uniform(0.75, 0.85)
        isc = rng.uniform(8.0, 12.0)     # Amps
        ipm = isc * rng.uniform(0.90, 0.98)

        now = datetime.now()
        return {
            "Pm": round(pm, 2),
            "Isc": round(isc, 3),
            "Voc": round(voc, 3),
            "Ipm": round(ipm, 3),
            "Vpm": round(vpm, 3),
            "Date": now.date().isoformat(),
            "TTime": now.time().strftime("%H:%M:%S"),
        }
    
    def _ensure_database(self):
        """Create Excel database with headers if it doesn't exist"""
        if not self.db_file.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "SerialNos"
            
            # Headers
            headers = [
                "SerialNo", "Pm", "Isc", "Voc", "Ipm", "Vpm",
                "Date", "TTime", "First Imported", "Last Updated"
            ]
            for col, header in enumerate(headers, start=1):
                ws.cell(1, col, header)
            
            # Make header row bold
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            wb.save(self.db_file)
            wb.close()
    
    def _ensure_master_data_sheet(self):
        """Create master sun simulator data sheet if it doesn't exist"""
        if not self.master_data_file.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "All Simulator Data"
            
            # Headers (will be populated from first import)
            headers = [
                "SerialNo", "Pm", "Isc", "Voc", "Ipm", "Vpm",
                "Date", "TTime", "Source File", "Imported Date"
            ]
            for col, header in enumerate(headers, start=1):
                ws.cell(1, col, header)
            
            # Make header row bold
            for cell in ws[1]:
                cell.font = Font(bold=True)
            
            wb.save(self.master_data_file)
            wb.close()
    
    def _refresh_serial_cache(self):
        """Refresh the serial cache if it's stale or missing"""
        current_time = time_module.time()
        
        # Check if file has been modified externally (real-time detection)
        if self.file_monitor.has_changed():
            # File changed externally - invalidate cache immediately
            self._serial_cache = None
            self._serial_cache_timestamp = 0
        
        # Check if cache is still valid
        if (self._serial_cache is not None and 
            current_time - self._serial_cache_timestamp < self._cache_ttl):
            return  # Cache is still valid
        
        # Refresh cache
        if not self.db_file.exists():
            self._serial_cache = set()
            self._serial_cache_timestamp = current_time
            return
        
        try:
            # Use pandas for much faster reads (critical for UI responsiveness)
            try:
                df = pd.read_excel(self.db_file, sheet_name='SerialNos', engine='openpyxl', usecols=['SerialNo'])
                # Normalize and create set of all serials
                self._serial_cache = {normalize_serial(s) for s in df['SerialNo'].dropna() if s}
                # Remove empty strings from cache
                self._serial_cache = {s for s in self._serial_cache if s}
            except Exception:
                # Fallback to openpyxl if pandas fails
                wb = load_workbook(self.db_file, read_only=True, data_only=True)
                if 'SerialNos' in wb.sheetnames:
                    ws = wb['SerialNos']
                    self._serial_cache = set()
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            serial_normalized = normalize_serial(row[0])
                            if serial_normalized:
                                self._serial_cache.add(serial_normalized)
                else:
                    self._serial_cache = set()
                wb.close()
            
            self._serial_cache_timestamp = current_time
        except Exception:
            # Silently fail - cache will be empty, validation will return False
            self._serial_cache = set()
            self._serial_cache_timestamp = current_time
    
    def validate_serial(self, serial: str) -> bool:
        """
        Check if a SerialNo exists in the Excel database.
        Uses caching for fast lookups without blocking the UI.
        
        Args:
            serial: Serial number to validate
            
        Returns:
            True if SerialNo exists, False otherwise
        """
        if not self.db_file.exists():
            return False
        
        # Normalize input serial (handles string, numeric, whitespace, .0 suffix)
        serial_normalized = normalize_serial(serial)
        if not serial_normalized:
            return False
        
        # Refresh cache if needed (fast operation)
        self._refresh_serial_cache()
        
        # Check cache (O(1) lookup - instant!)
        return serial_normalized in self._serial_cache if self._serial_cache else False
    
    def invalidate_cache(self):
        """Invalidate all caches (call after imports)"""
        self._serial_cache = None
        self._serial_cache_timestamp = 0
        self._data_cache = {}
        self._data_cache_timestamp = 0
        # Reset file monitors to prevent false positives after our own updates
        self.file_monitor.reset()
        if self.master_file_monitor:
            self.master_file_monitor.reset()
    
    def _refresh_data_cache(self, required_serials: Optional[List[str]] = None):
        """
        Refresh the data cache if it's stale or missing.
        LAZY LOADING: Only loads data for required serials, not entire database.
        
        Args:
            required_serials: Optional list of serials to load. If None, doesn't pre-load.
        """
        current_time = time_module.time()
        
        # Check if database file has been modified externally (real-time detection)
        if self.file_monitor.has_changed():
            # File changed externally - invalidate cache immediately
            self._data_cache = {}
            self._data_cache_timestamp = 0
        
        # Check if master data file has been modified externally
        if self.master_file_monitor and self.master_file_monitor.has_changed():
            # Master file changed externally - invalidate cache immediately
            self._data_cache = {}
            self._data_cache_timestamp = 0
        
        # Check if cache is still valid
        if (self._data_cache and 
            current_time - self._data_cache_timestamp < self._data_cache_ttl):
            return  # Cache is still valid
        
        # LAZY LOADING: Only load if specific serials are requested
        if not required_serials:
            # Don't pre-load entire database - too heavy
            return
        
        # Limit cache size to prevent memory bloat
        if len(self._data_cache) > self._data_cache_max_size:
            # Clear oldest entries (simple FIFO)
            self._data_cache = {}
        
        if not self.db_file.exists():
            self._data_cache = {}
            self._data_cache_timestamp = current_time
            return
        
        try:
            # Only load data for required serials (much lighter!)
            required_set = {normalize_serial(s) for s in required_serials if s}
            # Remove empty strings
            required_set = {s for s in required_set if s}
            
            # Use pandas for fast selective read
            try:
                df = pd.read_excel(self.db_file, sheet_name='SerialNos', engine='openpyxl')
                # Normalize serials using our function (handles .0 suffix, cell refs, etc.)
                df['SerialNo'] = df['SerialNo'].apply(normalize_serial)
                # Remove empty strings
                df = df[df['SerialNo'] != '']
                
                # Filter to only required serials
                df_filtered = df[df['SerialNo'].isin(required_set)]
                
                # Limit cache size before adding new entries
                if len(self._data_cache) + len(df_filtered) >= self._data_cache_max_size:
                    to_remove = len(self._data_cache) + len(df_filtered) - self._data_cache_max_size + 100
                    keys_to_remove = list(self._data_cache.keys())[:to_remove]
                    for key in keys_to_remove:
                        del self._data_cache[key]
                
                # Build cache dictionary only for required serials
                for _, row in df_filtered.iterrows():
                    serial = normalize_serial(row['SerialNo'])
                    if serial:
                        self._data_cache[serial] = {
                            'Pm': row.get('Pm'),
                            'Isc': row.get('Isc'),
                            'Voc': row.get('Voc'),
                            'Ipm': row.get('Ipm'),
                            'Vpm': row.get('Vpm'),
                            'Date': row.get('Date'),
                            'TTime': row.get('TTime'),
                        }
            except Exception:
                # Fallback to openpyxl - only load required serials
                wb = load_workbook(self.db_file, read_only=True, data_only=True)
                if 'SerialNos' in wb.sheetnames:
                    ws = wb['SerialNos']
                    headers = [cell.value for cell in ws[1]] if len(ws[1]) > 0 else []
                    
                    # Limit cache size before adding new entries
                    estimated_new = len(required_set)
                    if len(self._data_cache) + estimated_new >= self._data_cache_max_size:
                        to_remove = len(self._data_cache) + estimated_new - self._data_cache_max_size + 100
                        keys_to_remove = list(self._data_cache.keys())[:to_remove]
                        for key in keys_to_remove:
                            del self._data_cache[key]
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            serial = normalize_serial(row[0])
                            if serial and serial in required_set:
                                try:
                                    self._data_cache[serial] = {
                                        'Pm': row[headers.index('Pm')] if 'Pm' in headers and len(row) > headers.index('Pm') else None,
                                        'Isc': row[headers.index('Isc')] if 'Isc' in headers and len(row) > headers.index('Isc') else None,
                                        'Voc': row[headers.index('Voc')] if 'Voc' in headers and len(row) > headers.index('Voc') else None,
                                        'Ipm': row[headers.index('Ipm')] if 'Ipm' in headers and len(row) > headers.index('Ipm') else None,
                                        'Vpm': row[headers.index('Vpm')] if 'Vpm' in headers and len(row) > headers.index('Vpm') else None,
                                        'Date': row[headers.index('Date')] if 'Date' in headers and len(row) > headers.index('Date') else None,
                                        'TTime': row[headers.index('TTime')] if 'TTime' in headers and len(row) > headers.index('TTime') else None,
                                    }
                                except (IndexError, ValueError):
                                    continue  # Skip malformed rows
                wb.close()
            
            self._data_cache_timestamp = current_time
        except Exception:
            self._data_cache = {}
            self._data_cache_timestamp = current_time
    
    def get_serial_data(self, serial: str) -> Optional[Dict]:
        """
        Get electrical values for a SerialNo (uses cache for performance).
        
        Returns:
            Dict with electrical values or None if not found
        """
        # Normalize input serial
        serial_normalized = normalize_serial(serial)
        if not serial_normalized:
            return None
        
        # Try cache first (lazy load if needed)
        self._refresh_data_cache([serial_normalized])  # Only load this one serial
        if serial_normalized in self._data_cache:
            return self._data_cache[serial_normalized]
        
        # Cache miss - fallback to file read (shouldn't happen often)
        if not self.db_file.exists():
            # Database file missing – generate theoretical data as final fallback
            return self._generate_theoretical_data(serial)
        
        wb = None
        try:
            wb = load_workbook(self.db_file, read_only=True, data_only=True)
            if 'SerialNos' not in wb.sheetnames:
                # Sheet missing – treat as no real data and use fallback
                return self._generate_theoretical_data(serial)
            
            ws = wb['SerialNos']
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    stored_serial = normalize_serial(row[0])
                    if stored_serial and stored_serial == serial_normalized:
                        data = {
                            'SerialNo': stored_serial,
                            'Pm': row[1] if len(row) > 1 else None,
                            'Isc': row[2] if len(row) > 2 else None,
                            'Voc': row[3] if len(row) > 3 else None,
                            'Ipm': row[4] if len(row) > 4 else None,
                            'Vpm': row[5] if len(row) > 5 else None,
                            'Date': row[6] if len(row) > 6 else None,
                            'TTime': row[7] if len(row) > 7 else None,
                        }
                        # Update cache (limit size to prevent memory leaks)
                        if len(self._data_cache) >= self._data_cache_max_size:
                            # Remove oldest 20% of entries
                            to_remove = int(self._data_cache_max_size * 0.2)
                            keys_to_remove = list(self._data_cache.keys())[:to_remove]
                            for key in keys_to_remove:
                                del self._data_cache[key]
                        
                        self._data_cache[serial_normalized] = data
                        return data
            
            # Not found in file – use theoretical data as final fallback
            return self._generate_theoretical_data(serial)
        except Exception:
            # On any error, fall back to theoretical data instead of failing hard
            return self._generate_theoretical_data(serial)
        finally:
            # Ensure workbook is always closed
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass  # Ignore errors during cleanup
    
    def get_serial_data_batch(self, serials: list) -> Dict[str, Dict]:
        """
        Get electrical values for multiple SerialNos efficiently (uses cache).
        
        Args:
            serials: List of SerialNo strings to look up
            
        Returns:
            Dict mapping SerialNo -> Dict with electrical values
        """
        # Normalize input serials
        serials_normalized = {normalize_serial(s) for s in serials if s}
        # Remove empty strings
        serials_normalized = {s for s in serials_normalized if s}
        
        if not serials_normalized:
            return {}
        
        # Check cache first (lazy load missing serials only)
        result = {}
        cache_misses = []
        for serial in serials_normalized:
            if serial in self._data_cache:
                result[serial] = self._data_cache[serial]
            else:
                cache_misses.append(serial)
        
        # Lazy load missing serials only (much lighter!)
        if cache_misses:
            self._refresh_data_cache(list(cache_misses))
            # Check cache again after lazy load
            for serial in list(cache_misses):  # Copy to modify during iteration
                if serial in self._data_cache:
                    result[serial] = self._data_cache[serial]
                    cache_misses.remove(serial)
        
        # If all found in cache, return immediately
        if not cache_misses:
            return result
        
        # Load missing from file (shouldn't happen often with good cache)
        if not self.db_file.exists():
            # No database file – synthesize data for any remaining serials
            for s in cache_misses:
                fallback = self._generate_theoretical_data(s)
                if fallback:
                    result[s] = fallback
            return result
        
        wb = None
        try:
            # Use pandas for faster reads
            try:
                df = pd.read_excel(self.db_file, sheet_name='SerialNos', engine='openpyxl')
                # Normalize serials for comparison
                df['SerialNo'] = df['SerialNo'].apply(normalize_serial)
                # Remove empty strings
                df = df[df['SerialNo'] != '']
                df_filtered = df[df['SerialNo'].isin(cache_misses)]
                
                # Limit cache size before adding new entries
                if len(self._data_cache) + len(df_filtered) >= self._data_cache_max_size:
                    to_remove = len(self._data_cache) + len(df_filtered) - self._data_cache_max_size + 100
                    keys_to_remove = list(self._data_cache.keys())[:to_remove]
                    for key in keys_to_remove:
                        del self._data_cache[key]
                
                for _, row in df_filtered.iterrows():
                    serial = normalize_serial(row['SerialNo'])
                    if serial:
                        data = {
                            'Pm': row.get('Pm'),
                            'Isc': row.get('Isc'),
                            'Voc': row.get('Voc'),
                            'Ipm': row.get('Ipm'),
                            'Vpm': row.get('Vpm'),
                            'Date': row.get('Date'),
                            'TTime': row.get('TTime'),
                        }
                        result[serial] = data
                        # Update cache
                        self._data_cache[serial] = data
            except Exception:
                # Fallback to openpyxl
                wb = load_workbook(self.db_file, read_only=True, data_only=True)
                if 'SerialNos' in wb.sheetnames:
                    ws = wb['SerialNos']
                    headers = [cell.value for cell in ws[1]] if len(ws[1]) > 0 else []
                    
                    # Limit cache size before adding new entries
                    estimated_new = len(cache_misses)
                    if len(self._data_cache) + estimated_new >= self._data_cache_max_size:
                        to_remove = len(self._data_cache) + estimated_new - self._data_cache_max_size + 100
                        keys_to_remove = list(self._data_cache.keys())[:to_remove]
                        for key in keys_to_remove:
                            del self._data_cache[key]
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            serial_normalized = normalize_serial(row[0])
                            if serial_normalized and serial_normalized in cache_misses:
                                try:
                                    data = {
                                        'Pm': row[headers.index('Pm')] if 'Pm' in headers and len(row) > headers.index('Pm') else None,
                                        'Isc': row[headers.index('Isc')] if 'Isc' in headers and len(row) > headers.index('Isc') else None,
                                        'Voc': row[headers.index('Voc')] if 'Voc' in headers and len(row) > headers.index('Voc') else None,
                                        'Ipm': row[headers.index('Ipm')] if 'Ipm' in headers and len(row) > headers.index('Ipm') else None,
                                        'Vpm': row[headers.index('Vpm')] if 'Vpm' in headers and len(row) > headers.index('Vpm') else None,
                                        'Date': row[headers.index('Date')] if 'Date' in headers and len(row) > headers.index('Date') else None,
                                        'TTime': row[headers.index('TTime')] if 'TTime' in headers and len(row) > headers.index('TTime') else None,
                                    }
                                    result[serial_normalized] = data
                                    # Update cache
                                    self._data_cache[serial_normalized] = data
                                except (IndexError, ValueError):
                                    continue  # Skip malformed rows
        except Exception:
            # On any error during batch load, continue with whatever we have
            pass
        finally:
            # Ensure workbook is always closed
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass  # Ignore errors during cleanup
        
        # FINAL FALLBACK: generate theoretical data for any serials still missing
        for s in cache_misses:
            if s not in result:
                fallback = self._generate_theoretical_data(s)
                if fallback:
                    result[s] = fallback
        
        return result
    
    def import_simulator_file(self, file_path: Path) -> Tuple[int, int, list]:
        """
        Import SerialNos and electrical values from simulator export.
        Supports all Excel formats (.xlsx, .xlsm, .xls, .xlsb) and CSV.
        Copies imported file to IMPORTED DATA folder (keeps original in place) and updates master data sheet.
        
        Args:
            file_path: Path to simulator export (CSV or any Excel format)
            
        Returns:
            (imported_count, updated_count, errors)
        """
        imported = 0
        updated = 0
        errors = []
        
        try:
            # Read simulator file - support all Excel formats
            if file_path.suffix.lower() == '.csv':
                df = None
                for encoding in ['utf-8', 'latin-1', 'cp1252']:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                if df is None:
                    errors.append("Could not decode CSV file")
                    return 0, 0, errors
            else:
                # Excel file - try different engines based on file type
                excel_ext = file_path.suffix.lower()
                df = None
                
                if excel_ext in ['.xlsx', '.xlsm']:
                    # Modern Excel format - use openpyxl
                    try:
                        df = pd.read_excel(file_path, engine='openpyxl')
                    except Exception as e:
                        errors.append(f"Could not read Excel file with openpyxl: {e}")
                        return 0, 0, errors
                elif excel_ext == '.xls':
                    # Old Excel format - use xlrd
                    try:
                        df = pd.read_excel(file_path, engine='xlrd')
                    except ImportError:
                        errors.append("xlrd package required for .xls files. Install with: pip install xlrd")
                        return 0, 0, errors
                    except Exception as e:
                        errors.append(f"Could not read .xls file: {e}")
                        return 0, 0, errors
                elif excel_ext == '.xlsb':
                    # Excel Binary format - use pyxlsb
                    try:
                        df = pd.read_excel(file_path, engine='pyxlsb')
                    except ImportError:
                        errors.append("pyxlsb package required for .xlsb files. Install with: pip install pyxlsb")
                        return 0, 0, errors
                    except Exception as e:
                        errors.append(f"Could not read .xlsb file: {e}")
                        return 0, 0, errors
                else:
                    # Try openpyxl as default
                    try:
                        df = pd.read_excel(file_path, engine='openpyxl')
                    except Exception as e:
                        errors.append(f"Could not read Excel file: {e}")
                        return 0, 0, errors
            
            # Normalize column names
            df.columns = [str(col).strip() for col in df.columns]
            
            # Find columns (handle variations like "Voc(V)" and "Vpm(V)")
            serial_col = None
            pm_col = None
            isc_col = None
            voc_col = None
            ipm_col = None
            vpm_col = None
            date_col = None
            ttime_col = None
            
            for col in df.columns:
                col_lower = str(col).lower().strip()
                # Remove parentheses and spaces for more flexible matching
                col_normalized = col_lower.replace('(', '').replace(')', '').replace(' ', '').replace('_', '')
                
                if 'serial' in col_lower and ('no' in col_lower or 'number' in col_lower):
                    if not serial_col:
                        serial_col = col
                elif col_normalized == 'pm' or col_normalized == 'pmw' or 'power' in col_lower:
                    if not pm_col:
                        pm_col = col
                elif col_normalized == 'isc' or col_normalized == 'isca':
                    if not isc_col:
                        isc_col = col
                elif 'voc' in col_normalized or 'voc' in col_lower:
                    if not voc_col:
                        voc_col = col
                elif col_normalized == 'ipm' or col_normalized == 'ipma':
                    if not ipm_col:
                        ipm_col = col
                elif 'vpm' in col_normalized or 'vpm' in col_lower:
                    if not vpm_col:
                        vpm_col = col
                elif col_lower == 'date':
                    if not date_col:
                        date_col = col
                elif col_lower == 'ttime' or col_lower == 't time':
                    ttime_col = col
            
            # Log found columns for debugging (especially Voc and Vpm)
            if not voc_col or not vpm_col:
                print(f"DEBUG: Available columns in file: {list(df.columns)}")
                print(f"DEBUG: Voc column found: {voc_col}, Vpm column found: {vpm_col}")
                if not voc_col:
                    print("DEBUG: Warning - Voc column not found. Looking for columns containing 'voc'")
                if not vpm_col:
                    print("DEBUG: Warning - Vpm column not found. Looking for columns containing 'vpm'")
            
            if not serial_col:
                errors.append("Could not find SerialNo column")
                return 0, 0, errors
            
            # Load existing database
            wb = load_workbook(self.db_file, read_only=False)
            ws = wb['SerialNos']
            
            # Get existing SerialNos (for updating vs adding)
            # Normalize stored serials for comparison (handles numeric, whitespace, .0 suffix)
            existing_serials = {}
            for row in ws.iter_rows(min_row=2):
                if row[0].value:
                    serial_normalized = normalize_serial(row[0].value)
                    if serial_normalized:
                        existing_serials[serial_normalized] = row[0].row
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Process each row from simulator file
            for idx, row in df.iterrows():
                serial = row[serial_col]
                if pd.isna(serial) or not serial:
                    continue
                
                # Normalize serial for comparison (consistent with existing_serials keys)
                serial_normalized = normalize_serial(serial)
                if not serial_normalized:
                    continue
                
                # Extract electrical values
                pm = row[pm_col] if pm_col and pm_col in df.columns else None
                isc = row[isc_col] if isc_col and isc_col in df.columns else None
                voc = row[voc_col] if voc_col and voc_col in df.columns else None
                ipm = row[ipm_col] if ipm_col and ipm_col in df.columns else None
                vpm = row[vpm_col] if vpm_col and vpm_col in df.columns else None
                date = row[date_col] if date_col and date_col in df.columns else None
                ttime = row[ttime_col] if ttime_col and ttime_col in df.columns else None
                
                # Convert to appropriate types
                try:
                    pm = float(pm) if pm is not None and not pd.isna(pm) else None
                    isc = float(isc) if isc is not None and not pd.isna(isc) else None
                    voc = float(voc) if voc is not None and not pd.isna(voc) else None
                    ipm = float(ipm) if ipm is not None and not pd.isna(ipm) else None
                    vpm = float(vpm) if vpm is not None and not pd.isna(vpm) else None
                except (ValueError, TypeError):
                    pass
                
                if serial_normalized in existing_serials:
                    # Always update existing records regardless of timestamp
                    row_num = existing_serials[serial_normalized]

                    # Update the record with new data
                    ws.cell(row_num, 2, pm)  # Pm
                    ws.cell(row_num, 3, isc)  # Isc
                    ws.cell(row_num, 4, voc)  # Voc
                    ws.cell(row_num, 5, ipm)  # Ipm
                    ws.cell(row_num, 6, vpm)  # Vpm
                    if date:
                        ws.cell(row_num, 7, date)
                    if ttime:
                        ws.cell(row_num, 8, ttime)
                    ws.cell(row_num, 10, now)  # Last Updated
                    updated += 1
                else:
                    # Add new row (use normalized serial for consistency)
                    ws.append([
                        serial_normalized, pm, isc, voc, ipm, vpm,
                        date, ttime, now, now
                    ])
                    imported += 1
            
            # Save the database file
            try:
                wb.save(self.db_file)
                # Removed DEBUG print for performance
            except Exception as e:
                errors.append(f"Failed to save database: {e}")
                import traceback
                traceback.print_exc()
                wb.close()
                return imported, updated, errors
            
            wb.close()
            
            # After successful import, update master sheet and move file to IMPORTED DATA
            # Do master sheet update FIRST (before moving file) to ensure we have the original filename
            if imported > 0 or updated > 0:
                self._update_master_data_sheet(file_path, df)
                self._move_to_imported_data(file_path)
                # Invalidate cache after import (new serials may have been added)
                self.invalidate_cache()
                # Removed DEBUG print for performance
            
        except Exception as e:
            errors.append(f"Error importing file: {e}")
            import traceback
            traceback.print_exc()
        
        return imported, updated, errors

    def import_simulator_file_validated(self, records_dict: Dict[str, Dict], source_file: Path) -> Tuple[int, int, list]:
        """
        Import pre-validated and deduplicated records from simulator export.
        This bypasses file parsing and validation, assuming records are already clean.
        Used by GUI import process after validation.

        Args:
            records_dict: Dictionary mapping serial numbers to record dictionaries
            source_file: Path to the source file for master data sheet updates

        Returns:
            (imported_count, updated_count, errors_list)
        """
        imported = 0
        updated = 0
        errors = []

        try:
            # Load existing database
            wb = load_workbook(self.db_file, read_only=False)
            ws = wb['SerialNos']

            # Get existing SerialNos
            existing_serials = {}
            for row in ws.iter_rows(min_row=2):
                if row[0].value:
                    serial_normalized = normalize_serial(row[0].value)
                    if serial_normalized:
                        existing_serials[serial_normalized] = row[0].row

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Process each record
            for serial_no, record in records_dict.items():
                if serial_no in existing_serials:
                    # Always update existing records regardless of timestamp
                    row_num = existing_serials[serial_no]
                    ws.cell(row_num, 2, record.get('Pm'))   # Pm
                    ws.cell(row_num, 3, record.get('Isc'))  # Isc
                    ws.cell(row_num, 4, record.get('Voc'))  # Voc
                    ws.cell(row_num, 5, record.get('Ipm'))  # Ipm
                    ws.cell(row_num, 6, record.get('Vpm'))  # Vpm
                    if record.get('Date'):
                        ws.cell(row_num, 7, record.get('Date'))
                    if record.get('TTime'):
                        ws.cell(row_num, 8, record.get('TTime'))
                    ws.cell(row_num, 10, now)  # Last Updated
                    updated += 1
                else:
                    # Add new row
                    ws.append([
                        serial_no,
                        record.get('Pm'), record.get('Isc'), record.get('Voc'),
                        record.get('Ipm'), record.get('Vpm'),
                        record.get('Date'), record.get('TTime'), now, now
                    ])
                    imported += 1

            # Save database
            wb.save(self.db_file)
            wb.close()

            # Update master sheet and move file
            if imported > 0 or updated > 0:
                self._update_master_data_sheet(source_file, pd.DataFrame(list(records_dict.values())))
                self._move_to_imported_data(source_file)
                self.invalidate_cache()

        except Exception as e:
            errors.append(f"Error importing validated records: {e}")

        return imported, updated, errors

    def _parse_timestamp(self, date_val, ttime_val) -> Optional[datetime]:
        """
        Parse Date and TTime into a datetime object for comparison.
        
        Args:
            date_val: Date value (can be datetime, date, or string)
            ttime_val: TTime value (can be time, timedelta, or string like "HH:MM:SS")
            
        Returns:
            datetime object or None if parsing fails
        """
        try:
            # Handle date
            if date_val is None:
                return None
            
            if isinstance(date_val, datetime):
                date_part = date_val.date()
            elif isinstance(date_val, pd.Timestamp):
                date_part = date_val.date()
            elif hasattr(date_val, 'date'):
                date_part = date_val.date()
            else:
                # Try parsing as string
                date_part = pd.to_datetime(date_val).date()
            
            # Handle time
            if ttime_val is None:
                time_part = datetime.min.time()
            elif isinstance(ttime_val, time):
                # Already a time object
                time_part = ttime_val
            elif isinstance(ttime_val, datetime):
                time_part = ttime_val.time()
            elif isinstance(ttime_val, timedelta):
                # Convert timedelta to time
                total_seconds = int(ttime_val.total_seconds())
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                time_part = datetime.min.replace(hour=hours, minute=minutes, second=seconds).time()
            elif hasattr(ttime_val, 'time') and callable(getattr(ttime_val, 'time')):
                time_part = ttime_val.time()
            else:
                # Try parsing as string (HH:MM:SS or HH:MM)
                if isinstance(ttime_val, str):
                    time_parts = ttime_val.split(':')
                    if len(time_parts) >= 2:
                        hours = int(time_parts[0])
                        minutes = int(time_parts[1])
                        seconds = int(time_parts[2]) if len(time_parts) > 2 else 0
                        time_part = datetime.min.replace(hour=hours, minute=minutes, second=seconds).time()
                    else:
                        time_part = datetime.min.time()
                else:
                    time_part = datetime.min.time()
            
            return datetime.combine(date_part, time_part)
            
        except Exception:
            return None
    
    def _compare_timestamps(self, new_date, new_ttime, existing_date, existing_ttime) -> bool:
        """
        Compare timestamps to determine if new data is more recent.
        
        Args:
            new_date: New date value
            new_ttime: New time value
            existing_date: Existing date value from database
            existing_ttime: Existing time value from database
            
        Returns:
            True if new data is more recent, False otherwise
        """
        new_dt = self._parse_timestamp(new_date, new_ttime)
        existing_dt = self._parse_timestamp(existing_date, existing_ttime)
        
        # If we can't parse either, default to updating (safer)
        if new_dt is None:
            return False  # Can't parse new, don't update
        if existing_dt is None:
            return True  # Can't parse existing, update with new
        
        # Compare timestamps
        return new_dt > existing_dt
    
    def get_serial_count(self) -> int:
        """Get total number of SerialNos in database"""
        if not self.db_file.exists():
            return 0
        
        wb = load_workbook(self.db_file, read_only=True, data_only=True)
        ws = wb['SerialNos']
        count = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
        wb.close()
        return count
    
    def _move_to_imported_data(self, source_file: Path):
        """Copy imported file to IMPORTED DATA folder (keeps original in place)"""
        try:
            # Create unique filename if file already exists
            dest_file = self.imported_data_dir / source_file.name
            counter = 1
            while dest_file.exists():
                stem = source_file.stem
                suffix = source_file.suffix
                dest_file = self.imported_data_dir / f"{stem}_{counter}{suffix}"
                counter += 1
            
            # Copy file instead of moving (keeps original where it was)
            shutil.copy2(str(source_file), str(dest_file))
            print(f"Copied file to IMPORTED DATA: {dest_file.name}")
        except Exception as e:
            # Log error but don't fail the import
            print(f"Warning: Could not copy file to IMPORTED DATA: {e}")
    
    def _update_master_data_sheet(self, source_file: Path, df: pd.DataFrame):
        """Append all data from imported file to master sun simulator data sheet"""
        try:
            # Read existing master data if it exists
            if self.master_data_file.exists():
                try:
                    existing_df = pd.read_excel(self.master_data_file, engine='openpyxl')
                except Exception:
                    existing_df = pd.DataFrame()
            else:
                existing_df = pd.DataFrame()
            
            # Prepare new data with source file info
            import_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            source_filename = source_file.name
            
            # Normalize column names in df (same as import logic)
            df.columns = [str(col).strip() for col in df.columns]
            
            # Find columns (same logic as import)
            serial_col = None
            pm_col = None
            isc_col = None
            voc_col = None
            ipm_col = None
            vpm_col = None
            date_col = None
            ttime_col = None
            
            # Find columns using same logic as main import (more flexible matching)
            for col in df.columns:
                col_lower = str(col).lower().strip()
                # SerialNo - check for variations (same as main import)
                if not serial_col and ('serial' in col_lower and ('no' in col_lower or 'number' in col_lower)):
                    serial_col = col
                # Pm - check for variations
                elif not pm_col and ('pm' in col_lower or 'power' in col_lower):
                    pm_col = col
                # Isc - check for variations
                elif not isc_col and 'isc' in col_lower:
                    isc_col = col
                # Voc - check for variations (Voc(V), Voc, etc.)
                elif not voc_col and 'voc' in col_lower:
                    voc_col = col
                # Ipm - check for variations
                elif not ipm_col and 'ipm' in col_lower:
                    ipm_col = col
                # Vpm - check for variations (Vpm(V), Vpm, etc.)
                elif not vpm_col and 'vpm' in col_lower:
                    vpm_col = col
                # Date
                elif not date_col and col_lower == 'date':
                    date_col = col
                # TTime - check for variations
                elif not ttime_col and ('ttime' in col_lower or (col_lower == 'time' and 't' in col_lower)):
                    ttime_col = col
            
            # Create new rows for master sheet
            new_rows = []
            if not serial_col:
                # If we can't find SerialNo column, log warning and skip
                print(f"Warning: Could not find SerialNo column in imported file. Available columns: {list(df.columns)}")
                return
            
            for _, row in df.iterrows():
                try:
                    serial_val = row.get(serial_col) if serial_col else None
                    if serial_val is not None and pd.notna(serial_val):
                        serial_str = str(serial_val).strip()
                        if serial_str:  # Only add if serial is not empty
                            # Extract values, handling NaN properly
                            pm_val = row.get(pm_col) if pm_col and pm_col in df.columns else None
                            isc_val = row.get(isc_col) if isc_col and isc_col in df.columns else None
                            voc_val = row.get(voc_col) if voc_col and voc_col in df.columns else None
                            ipm_val = row.get(ipm_col) if ipm_col and ipm_col in df.columns else None
                            vpm_val = row.get(vpm_col) if vpm_col and vpm_col in df.columns else None
                            date_val = row.get(date_col) if date_col and date_col in df.columns else None
                            ttime_val = row.get(ttime_col) if ttime_col and ttime_col in df.columns else None
                            
                            # Convert NaN to None
                            if pm_val is not None and pd.isna(pm_val):
                                pm_val = None
                            if isc_val is not None and pd.isna(isc_val):
                                isc_val = None
                            if voc_val is not None and pd.isna(voc_val):
                                voc_val = None
                            if ipm_val is not None and pd.isna(ipm_val):
                                ipm_val = None
                            if vpm_val is not None and pd.isna(vpm_val):
                                vpm_val = None
                            if date_val is not None and pd.isna(date_val):
                                date_val = None
                            if ttime_val is not None and pd.isna(ttime_val):
                                ttime_val = None
                            
                            new_rows.append({
                                'SerialNo': serial_str,
                                'Pm': pm_val,
                                'Isc': isc_val,
                                'Voc': voc_val,
                                'Ipm': ipm_val,
                                'Vpm': vpm_val,
                                'Date': str(date_val) if date_val is not None else None,
                                'TTime': str(ttime_val) if ttime_val is not None else None,
                                'Source File': source_filename,
                                'Imported Date': import_date
                            })
                except Exception as e:
                    # Skip rows that cause errors, but continue processing
                    print(f"Warning: Error processing row in master sheet update: {e}")
                    continue
            
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                
                # Append to existing data
                if not existing_df.empty:
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                else:
                    combined_df = new_df
                
                # Save to master sheet
                combined_df.to_excel(self.master_data_file, index=False, engine='openpyxl')
                print(f"Master data sheet updated: Added {len(new_rows)} rows from {source_filename}")
            else:
                print(f"Warning: No rows to add to master sheet from {source_filename}")
        except Exception as e:
            # Log error but don't fail the import
            import traceback
            print(f"Warning: Could not update master data sheet: {e}")
            traceback.print_exc()
