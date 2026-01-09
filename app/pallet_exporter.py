#!/usr/bin/env python3
"""
Pallet Exporter - Excel export functionality

Exports pallet-specific Excel files with the same format as the original workbook.
Creates standalone files ready for printing.
"""

from pathlib import Path
from datetime import datetime
from typing import Dict, Optional, Tuple
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Fill, Alignment, Protection
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import copy
import shutil


class PalletExporter:
    """Handles exporting pallets to Excel files"""
    
    def __init__(self, source_workbook: Path, export_dir: Path, serial_db=None):
        """
        Initialize PalletExporter.
        
        Args:
            source_workbook: Path to source pallet workbook
            export_dir: Base directory for exports (will create date subfolders)
            serial_db: Optional SerialDatabase instance for looking up electrical values
        """
        self.source_workbook = source_workbook
        self.base_export_dir = export_dir
        self.serial_db = serial_db
    
    def export_pallet(self, pallet: Dict, panel_type: Optional[str] = None, 
                     customer: Optional[Dict] = None,
                     progress_callback: Optional[callable] = None) -> Path:
        """
        Export a pallet to a standalone Excel file.
        
        Copies the entire reference workbook and renames it to the pallet number.
        Updates the PALLET SHEET with:
        - Cell B1 (panel type)
        - Cell B3 (panel type + date + pallet number)
        - Cell G3 (date)
        - Serial numbers and electrical values in the pallet slots
        Uses date-based folders and finds the next available pallet number to prevent overwrites.
        
        Args:
            pallet: Pallet dict with serial_numbers list
            panel_type: Panel type string (e.g., "200WT", "220WT", etc.) to write to Cell B1
            
        Returns:
            Path to exported Excel file
            
        Raises:
            FileNotFoundError: If source workbook doesn't exist
            KeyError: If required sheets are missing
            PermissionError: If source workbook is locked
        """
        if not self.source_workbook.exists():
            raise FileNotFoundError(f"Source workbook not found: {self.source_workbook}")
        
        # Progress: 0-10% - Setup
        if progress_callback:
            progress_callback("Preparing export...", 5)
        
        # Get date-based export directory (creates if needed)
        try:
            export_dir = self._get_export_dir()
        except RuntimeError as e:
            # Re-raise with more context about the export folder
            raise RuntimeError(
                f"Export folder error: {e}\n\n"
                f"Base export directory: {self.base_export_dir}\n"
                f"Please ensure this location exists and is writable."
            ) from e
        
        # Progress: 10-20% - Copying file
        if progress_callback:
            progress_callback("Copying workbook...", 15)
        
        # Create a temporary filename first (we'll determine final name after setting B3)
        temp_filename = "temp_pallet_export.xlsx"
        temp_export_path = export_dir / temp_filename
        
        # Copy the entire source workbook to the export location
        # Use copyfile instead of copy2 for faster operation (we don't need metadata)
        shutil.copyfile(self.source_workbook, temp_export_path)
        
        # Progress: 20-30% - Loading workbook
        if progress_callback:
            progress_callback("Loading workbook...", 25)
        
        # Open the copied workbook and update it
        # Optimized for older hardware: skip VBA, links, and data_only for faster loading
        try:
            wb = load_workbook(temp_export_path, read_only=False, keep_vba=False, keep_links=False, data_only=False)
        except PermissionError:
            raise PermissionError(f"Could not open workbook for editing: {temp_export_path}")
        
        try:
            # Validate pallet has serial numbers
            serials = pallet.get('serial_numbers', [])
            if not serials:
                raise ValueError("Cannot export empty pallet (no serial numbers)")
            
            # Progress: 30-60% - Updating PALLET SHEET
            if progress_callback:
                progress_callback("Updating pallet sheet...", 35)
            
            # Update PALLET SHEET if it exists (case-insensitive, handle variations)
            pallet_sheet_name = None
            for sheet_name in wb.sheetnames:
                if sheet_name.upper().replace(' ', '') == 'PALLETSHEET':
                    pallet_sheet_name = sheet_name
                    break
            
            if pallet_sheet_name:
                try:
                    # Update the sheet (this sets Cell B3)
                    self._update_pallet_sheet(wb[pallet_sheet_name], pallet, panel_type, customer, progress_callback)
                    if progress_callback:
                        progress_callback("Pallet sheet updated", 75)
                    
                    # Read Cell B3 value to use as filename
                    sheet = wb[pallet_sheet_name]
                    b3_cell = sheet['B3']
                    b3_value = str(b3_cell.value) if b3_cell.value else None
                    
                    if b3_value:
                        # Sanitize filename: remove invalid filesystem characters
                        # Windows: < > : " / \ | ? *
                        # macOS/Linux: / (and null)
                        invalid_chars = '<>:"/\\|?*'
                        sanitized_name = ''.join(c if c not in invalid_chars else '_' for c in b3_value)
                        filename = f"{sanitized_name}.xlsx"
                    else:
                        # Fallback to numbered filename if B3 is empty
                        available_number = self._find_next_available_pallet_number(export_dir)
                        filename = f"Pallet_{available_number}.xlsx"
                    
                    # Final export path using B3 value
                    export_path = export_dir / filename
                    
                    # Handle filename collision: if file already exists, append "_copy" or "_copy_2", etc.
                    if export_path.exists():
                        base_name = sanitized_name if b3_value else f"Pallet_{available_number}"
                        # First try "_copy"
                        filename = f"{base_name}_copy.xlsx"
                        export_path = export_dir / filename
                        counter = 2
                        # If "_copy" also exists, try "_copy_2", "_copy_3", etc.
                        while export_path.exists():
                            filename = f"{base_name}_copy_{counter}.xlsx"
                            export_path = export_dir / filename
                            counter += 1
                            if counter > 1000:  # Safety limit
                                raise RuntimeError("Too many files with same name - please clean up export directory")
                    
                except Exception as e:
                    raise RuntimeError(f"Failed to update PALLET SHEET: {e}") from e
            else:
                # PALLET SHEET is required - this is an error
                available_sheets = ', '.join(wb.sheetnames)
                raise KeyError(
                    f"PALLET SHEET not found in workbook.\n\n"
                    f"Available sheets: {available_sheets}\n\n"
                    f"Please ensure the reference workbook has a sheet named 'PALLET SHEET'."
                )
            
            # Progress: 60-80% - Preparing to save
            if progress_callback:
                progress_callback("Preparing to save...", 75)
            
            # Progress: 80-95% - Saving workbook
            if progress_callback:
                progress_callback("Saving workbook...", 85)
            
            # Save the updated workbook to the final export path (using B3 value as filename)
            # All formatting (column widths, row heights, merged cells, styles) is preserved
            try:
                wb.save(export_path)
                
                # Remove temp file if it exists and is different from final path
                if temp_export_path.exists() and temp_export_path != export_path:
                    try:
                        temp_export_path.unlink()
                    except Exception:
                        pass  # Non-critical if temp file cleanup fails
                
                if progress_callback:
                    progress_callback("Export complete!", 100)
            except PermissionError:
                raise PermissionError(f"Cannot save workbook - file may be open in Excel: {export_path}")
            except Exception as e:
                raise RuntimeError(f"Failed to save workbook: {e}") from e
            
            return export_path
            
        except (ValueError, RuntimeError, PermissionError):
            # Re-raise these specific errors
            raise
        except Exception as e:
            # Wrap unexpected errors
            raise RuntimeError(f"Unexpected error during export: {e}") from e
        finally:
            # Always close workbook
            try:
                if 'wb' in locals():
                    wb.close()
            except Exception:
                pass  # Ignore errors during close
            
            # Always clean up temp file if it exists and is different from export path
            try:
                if 'temp_export_path' in locals() and temp_export_path.exists():
                    # Only delete if it's actually a temp file (not the final export)
                    if 'export_path' in locals() and temp_export_path != export_path:
                        temp_export_path.unlink()
            except Exception:
                pass  # Non-critical if temp file cleanup fails
    
    def _get_export_dir(self) -> Path:
        """Get date-based export directory (creates if needed)"""
        # Ensure base export directory exists first
        try:
            self.base_export_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            raise RuntimeError(
                f"Cannot create export folder: {self.base_export_dir}\n"
                f"Error: {e}\n\n"
                f"Please ensure you have write permissions to this location."
            ) from e
        
        current_date = datetime.now()
        try:
            date_folder = current_date.strftime("%-d-%b-%y")  # Unix: removes leading zero
        except ValueError:
            date_folder = current_date.strftime("%#d-%b-%y")  # Windows: removes leading zero
        
        export_dir = self.base_export_dir / date_folder
        try:
            export_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            raise RuntimeError(
                f"Cannot create date-based export folder: {export_dir}\n"
                f"Error: {e}\n\n"
                f"Base folder: {self.base_export_dir}\n"
                f"Please ensure you have write permissions."
            ) from e
        
        return export_dir
    
    def _find_next_available_pallet_number(self, export_dir: Path) -> int:
        """
        Find the next available pallet number in the export directory.
        
        Scans existing Pallet_*.xlsx files and returns the next available number.
        If no files exist, returns 1.
        
        Args:
            export_dir: Directory to search for existing pallet files
            
        Returns:
            Next available pallet number
        """
        if not export_dir.exists():
            return 1
        
        # Find all existing pallet files
        existing_numbers = []
        for file in export_dir.glob("Pallet_*.xlsx"):
            # Extract number from filename like "Pallet_1.xlsx" or "Pallet_123.xlsx"
            try:
                # Remove "Pallet_" prefix and ".xlsx" suffix
                number_str = file.stem.replace("Pallet_", "")
                number = int(number_str)
                existing_numbers.append(number)
            except ValueError:
                # Skip files that don't match the pattern
                continue
        
        if not existing_numbers:
            return 1
        
        # Find the next available number
        max_number = max(existing_numbers)
        return max_number + 1
    
    def _validate_pm_range(self, pm_value: float, panel_type: Optional[str]) -> bool:
        """
        Validate that Pm value falls within expected range for panel type.
        
        Ranges:
        - 200WT: 195-206
        - 220WT: 214-227
        - 330WT: 320-340
        - 450WT: 439-463.5
        - 450BT: 439-463.5
        
        Returns True if valid, False otherwise.
        """
        if not panel_type or pm_value is None:
            return True  # Can't validate without panel type or value
        
        ranges = {
            '200WT': (195, 206),
            '220WT': (214, 227),
            '330WT': (320, 340),
            '450WT': (439, 463.5),
            '450BT': (439, 463.5),
        }
        
        if panel_type not in ranges:
            return True  # Unknown panel type, don't block
        
        min_val, max_val = ranges[panel_type]
        return min_val <= pm_value <= max_val
    
    def _copy_cell_style(self, source_cell, target_cell):
        """Copy all formatting from source cell to target cell"""
        if source_cell.has_style:
            # Copy font
            if source_cell.font:
                target_cell.font = Font(
                    name=source_cell.font.name,
                    size=source_cell.font.size,
                    bold=source_cell.font.bold,
                    italic=source_cell.font.italic,
                    vertAlign=source_cell.font.vertAlign,
                    underline=source_cell.font.underline,
                    strike=source_cell.font.strike,
                    color=source_cell.font.color
                )
            
            # Copy border
            if source_cell.border:
                target_cell.border = Border(
                    left=source_cell.border.left,
                    right=source_cell.border.right,
                    top=source_cell.border.top,
                    bottom=source_cell.border.bottom,
                    diagonal=source_cell.border.diagonal,
                    diagonal_direction=source_cell.border.diagonal_direction,
                    outline=source_cell.border.outline,
                    vertical=source_cell.border.vertical,
                    horizontal=source_cell.border.horizontal
                )
            
            # Copy fill
            if source_cell.fill:
                target_cell.fill = Fill(
                    fill_type=source_cell.fill.fill_type,
                    start_color=source_cell.fill.start_color,
                    end_color=source_cell.fill.end_color,
                    patternType=source_cell.fill.patternType
                )
            
            # Copy alignment
            if source_cell.alignment:
                target_cell.alignment = Alignment(
                    horizontal=source_cell.alignment.horizontal,
                    vertical=source_cell.alignment.vertical,
                    text_rotation=source_cell.alignment.text_rotation,
                    wrap_text=source_cell.alignment.wrap_text,
                    shrink_to_fit=source_cell.alignment.shrink_to_fit,
                    indent=source_cell.alignment.indent
                )
            
            # Copy number format
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            
            # Copy protection
            if source_cell.protection:
                target_cell.protection = Protection(
                    locked=source_cell.protection.locked,
                    hidden=source_cell.protection.hidden
                )
    
    def _update_pallet_sheet(self, sheet, pallet: Dict, panel_type: Optional[str] = None,
                            customer: Optional[Dict] = None,
                            progress_callback: Optional[callable] = None):
        """
        Update PALLET SHEET with panel type, date, serial numbers, and electrical values.
        
        Formatting is automatically preserved by openpyxl when we only modify .value.
        Column widths, row heights, merged cells, and cell styles are preserved
        because we use shutil.copy2() to copy the entire workbook first.
        """
        # Set customer information in Cell A3 (formatted with line breaks)
        if customer:
            customer_text = f"{customer['name']}\n{customer['business']}\n{customer['address']}\n{customer['city']}, {customer['state']} {customer['zip_code']}"
            # Update Cell A3 - handle merged cells
            try:
                cell_a3 = sheet.cell(row=3, column=1)  # A is column 1
                if not isinstance(cell_a3, MergedCell):
                    # Not merged, set directly (fastest)
                    cell_a3.value = customer_text
                else:
                    # Is merged, need to unmerge/remerge
                    merged_range_a3 = None
                    merge_range_obj_a3 = None
                    for merge_range in list(sheet.merged_cells.ranges):
                        if 'A3' in str(merge_range):
                            merged_range_a3 = str(merge_range)
                            merge_range_obj_a3 = merge_range
                            break
                    
                    if merged_range_a3:
                        sheet.unmerge_cells(merged_range_a3)
                        sheet.cell(row=merge_range_obj_a3.min_row, column=merge_range_obj_a3.min_col).value = customer_text
                        sheet.merge_cells(merged_range_a3)
                    else:
                        sheet.cell(row=3, column=1).value = customer_text
            except Exception:
                # Fallback: direct access
                sheet.cell(row=3, column=1).value = customer_text
        
        # Set panel type in Cell B1 (formatting automatically preserved)
        # Handle merged cells - if B1 is part of a merge, we need to unmerge, set value, then remerge
        # Optimized: Cache merged ranges to avoid repeated iteration
        if panel_type:
            # Quick check: try direct access first (fastest path)
            try:
                cell_b1 = sheet.cell(row=1, column=2)
                if not isinstance(cell_b1, MergedCell):
                    # Not merged, set directly (fastest)
                    cell_b1.value = panel_type
                else:
                    # Is merged, need to unmerge/remerge
                    merged_range_b1 = None
                    merge_range_obj_b1 = None
                    for merge_range in list(sheet.merged_cells.ranges):
                        if 'B1' in str(merge_range):
                            merged_range_b1 = str(merge_range)
                            merge_range_obj_b1 = merge_range
                            break
                    
                    if merged_range_b1:
                        sheet.unmerge_cells(merged_range_b1)
                        sheet.cell(row=merge_range_obj_b1.min_row, column=merge_range_obj_b1.min_col).value = panel_type
                        sheet.merge_cells(merged_range_b1)
                    else:
                        # Fallback: direct access
                        sheet.cell(row=1, column=2).value = panel_type
            except Exception:
                # Fallback: direct access if anything fails
                sheet.cell(row=1, column=2).value = panel_type
        
        # Calculate and set weight in Cell D2: number of panels × 40
        serials = pallet.get('serial_numbers', [])
        panel_count = len(serials) if serials else 0
        weight = panel_count * 40
        
        # Update Cell D2 with calculated weight - handle merged cells
        try:
            cell_d2 = sheet.cell(row=2, column=4)  # D is column 4
            if not isinstance(cell_d2, MergedCell):
                # Not merged, set directly (fastest)
                cell_d2.value = weight
            else:
                # Is merged, need to unmerge/remerge
                merged_range_d2 = None
                merge_range_obj_d2 = None
                for merge_range in list(sheet.merged_cells.ranges):
                    if 'D2' in str(merge_range):
                        merged_range_d2 = str(merge_range)
                        merge_range_obj_d2 = merge_range
                        break
                
                if merged_range_d2:
                    sheet.unmerge_cells(merged_range_d2)
                    sheet.cell(row=merge_range_obj_d2.min_row, column=merge_range_obj_d2.min_col).value = weight
                    sheet.merge_cells(merged_range_d2)
                else:
                    sheet.cell(row=2, column=4).value = weight
        except Exception:
            # Fallback: direct access
            sheet.cell(row=2, column=4).value = weight
        
        # Set current date in Cell G3 (formatted as d-Mmm-yy, e.g., "6-Jan-26")
        # Formatting automatically preserved
        current_date = datetime.now()
        try:
            date_formatted = current_date.strftime("%-d-%b-%y")  # Unix: removes leading zero
        except ValueError:
            date_formatted = current_date.strftime("%#d-%b-%y")  # Windows: removes leading zero
        # Update Cell G3 with date - handle merged cells (optimized)
        try:
            cell_g3 = sheet.cell(row=3, column=7)
            if not isinstance(cell_g3, MergedCell):
                # Not merged, set directly (fastest)
                cell_g3.value = date_formatted
            else:
                # Is merged, need to unmerge/remerge
                merged_range_g3 = None
                merge_range_obj_g3 = None
                for merge_range in list(sheet.merged_cells.ranges):
                    if 'G3' in str(merge_range):
                        merged_range_g3 = str(merge_range)
                        merge_range_obj_g3 = merge_range
                        break
                
                if merged_range_g3:
                    sheet.unmerge_cells(merged_range_g3)
                    sheet.cell(row=merge_range_obj_g3.min_row, column=merge_range_obj_g3.min_col).value = date_formatted
                    sheet.merge_cells(merged_range_g3)
                else:
                    sheet.cell(row=3, column=7).value = date_formatted
        except Exception:
            # Fallback: direct access
            sheet.cell(row=3, column=7).value = date_formatted
        
        # Set Cell B3 with format: {PanelType}{MDYYYY}-{PalletNumber}
        # Example: 200WT2192025-18 (200WT, Feb 19 2025, Pallet 18)
        #         220WT162025-1 (220WT, Jan 6 2025, Pallet 1)
        # Format: Single digit month (no leading zero), single digit day (no leading zero), four digit year
        if panel_type:
            pallet_number = pallet.get('pallet_number', 1)
            # Format date as MDYYYY (e.g., 2192025 for February 19, 2025, or 162025 for January 6, 2025)
            # Use %-m and %-d for single digits (Unix) or %#m and %#d (Windows)
            try:
                date_mdyyyy = current_date.strftime("%-m%-d%Y")  # Unix: removes leading zeros from month and day
            except ValueError:
                date_mdyyyy = current_date.strftime("%#m%#d%Y")  # Windows: removes leading zeros from month and day
            # Combine: PanelType + MDYYYY + "-" + PalletNumber
            b3_value = f"{panel_type}{date_mdyyyy}-{pallet_number}"
            # Update Cell B3 - handle merged cells (optimized)
            try:
                cell_b3 = sheet.cell(row=3, column=2)
                if not isinstance(cell_b3, MergedCell):
                    # Not merged, set directly (fastest)
                    cell_b3.value = b3_value
                else:
                    # Is merged, need to unmerge/remerge
                    merged_range_b3 = None
                    merge_range_obj_b3 = None
                    for merge_range in list(sheet.merged_cells.ranges):
                        if 'B3' in str(merge_range):
                            merged_range_b3 = str(merge_range)
                            merge_range_obj_b3 = merge_range
                            break
                    
                    if merged_range_b3:
                        sheet.unmerge_cells(merged_range_b3)
                        sheet.cell(row=merge_range_obj_b3.min_row, column=merge_range_obj_b3.min_col).value = b3_value
                        sheet.merge_cells(merged_range_b3)
                    else:
                        sheet.cell(row=3, column=2).value = b3_value
            except Exception:
                # Fallback: direct access
                sheet.cell(row=3, column=2).value = b3_value
        
        # Populate serial numbers and electrical values in PALLET SHEET
        serials = pallet.get('serial_numbers', [])
        
        # Serial numbers go in column B, rows 5-30
        serial_col = 'B'
        start_row = 5
        
        # Find columns for electrical values
        pm_col = self._find_column_by_header(sheet, ['Pm', 'Pm(W)', 'Pm (W)'])
        isc_col = self._find_column_by_header(sheet, ['Isc', 'Isc(A)', 'Isc (A)'])
        voc_col = self._find_column_by_header(sheet, ['Voc', 'Voc(V)', 'Voc (V)', 'Voc(V)'])
        ipm_col = self._find_column_by_header(sheet, ['Ipm', 'Ipm(A)', 'Ipm (A)'])
        vpm_col = self._find_column_by_header(sheet, ['Vpm', 'Vpm(V)', 'Vpm (V)', 'Vpm(V)'])
        
        # Populate serials and electrical values
        # Note: openpyxl automatically preserves cell formatting when we only change .value
        # Column widths, row heights, and merged cells are preserved by shutil.copy2()
        
        # Progress update: Loading electrical data
        if progress_callback:
            progress_callback("Loading electrical data...", 40)
        
        # Batch load all serial data at once to avoid repeated file I/O (prevents freeze)
        # This opens the database file ONCE instead of 25 times
        serial_data_cache = {}
        if self.serial_db and serials:
            try:
                # Use batch method to load all serial data in one file operation
                if hasattr(self.serial_db, 'get_serial_data_batch'):
                    serial_data_cache = self.serial_db.get_serial_data_batch(serials)
                else:
                    # Fallback to individual lookups if batch method doesn't exist
                    for serial in serials:
                        try:
                            serial_data = self.serial_db.get_serial_data(serial)
                            if serial_data:
                                serial_data_cache[serial] = serial_data
                        except Exception:
                            # Continue if individual lookup fails
                            continue
            except Exception as e:
                # If batch load fails, log but continue without electrical values
                # This allows export to complete even if database is temporarily unavailable
                print(f"Warning: Could not load electrical values from database: {e}")
                serial_data_cache = {}
        
        # Progress update: Populating cells
        if progress_callback:
            progress_callback("Populating cells...", 45)
        
        # Now populate cells using cached data
        total_serials = len(serials)
        for i, serial in enumerate(serials):
            row = start_row + i
            if row <= sheet.max_row if sheet.max_row else 100:
                # Set serial number (formatting preserved automatically)
                sheet[f'{serial_col}{row}'].value = serial
                
                # Use cached electrical values (much faster than individual lookups)
                serial_data = serial_data_cache.get(serial)
                if serial_data:
                    # Validate Pm value against panel type range (silent validation)
                    pm_value = serial_data.get('Pm')
                    if pm_value is not None:
                        is_valid = self._validate_pm_range(pm_value, panel_type)
                        # Still populate even if out of range (validation is for app knowledge only)
                        if pm_col:
                            sheet[f'{pm_col}{row}'].value = pm_value
                    
                    # Populate other electrical values (always use most recent from database)
                    # Formatting is automatically preserved when we only change .value
                    if isc_col and serial_data.get('Isc') is not None:
                        sheet[f'{isc_col}{row}'].value = serial_data['Isc']
                    if voc_col and serial_data.get('Voc') is not None:
                        sheet[f'{voc_col}{row}'].value = serial_data['Voc']
                    if ipm_col and serial_data.get('Ipm') is not None:
                        sheet[f'{ipm_col}{row}'].value = serial_data['Ipm']
                    if vpm_col and serial_data.get('Vpm') is not None:
                        sheet[f'{vpm_col}{row}'].value = serial_data['Vpm']
                
                # Update progress every 5 serials for smoother progress bar
                if progress_callback and (i + 1) % 5 == 0:
                    percent = 45 + int((i + 1) / total_serials * 5)  # 45-50% range
                    progress_callback(f"Populating cells... ({i + 1}/{total_serials})", percent)
    
    def _find_column_by_header(self, sheet, header_variations: list) -> Optional[str]:
        """Find column letter by searching for header text variations in rows 1-5"""
        # Optimized: use max_col to limit search, cache header_variations_lower
        header_variations_lower = [v.lower() for v in header_variations]
        max_col = min(sheet.max_column if hasattr(sheet, 'max_column') else 26, 26)  # Limit to Z
        
        for row in range(1, min(6, sheet.max_row + 1) if sheet.max_row else 6):
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                cell = sheet[f'{col_letter}{row}']
                if cell.value:
                    cell_value = str(cell.value).strip().lower()
                    for variation_lower in header_variations_lower:
                        if variation_lower in cell_value or cell_value in variation_lower:
                            return col_letter
        return None
    
    def _find_serial_column(self, sheet) -> str:
        """Find which column contains serial numbers by looking for VLOOKUP formulas"""
        # Common patterns: column A or B
        # Check first few rows for VLOOKUP formulas that reference SerialNo
        for row in range(1, min(10, sheet.max_row + 1) if sheet.max_row else 10):
            for col_letter in ['A', 'B', 'C']:
                cell = sheet[f'{col_letter}{row}']
                if cell.value and isinstance(cell.value, str):
                    # Check if it's a VLOOKUP formula referencing SerialNo
                    if 'VLOOKUP' in str(cell.value).upper() and 'SERIALNO' in str(cell.value).upper():
                        # Serial numbers are typically in column A or B
                        # If VLOOKUP is in column C+, serials are likely in A or B
                        return 'A'  # Default to A
        # Default to column A if we can't detect
        return 'A'
    
    def _find_serial_start_row(self, sheet, serial_col: str) -> int:
        """Find the starting row for serial numbers"""
        # Serial numbers go in cells B5 through B30
        return 5
    
