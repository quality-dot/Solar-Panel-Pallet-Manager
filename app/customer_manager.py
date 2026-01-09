#!/usr/bin/env python3
"""
Customer Manager - Manages customer information for pallet exports

Stores customer data in Excel format for easy manual editing.
"""

from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime, timedelta


class CustomerManager:
    """Manages customer information for pallet exports"""
    
    def __init__(self, excel_file: Path):
        """
        Initialize CustomerManager.
        
        Args:
            excel_file: Path to Excel file storing customer data
        """
        self.excel_file = excel_file
        self.customers: List[Dict] = []
        self._last_load_time: Optional[datetime] = None
        self._cache_ttl = timedelta(seconds=5)  # Cache customers for 5 seconds
        self._file_modified_time: Optional[datetime] = None
        self._load_customers()
    
    def _load_customers(self, force_reload=False):
        """
        Load customers from Excel file with caching for performance.
        
        Args:
            force_reload: If True, bypass cache and reload from file
        """
        # Check if we need to reload based on cache
        if not force_reload and self._last_load_time:
            try:
                # Check if file modification time has changed
                current_mtime = datetime.fromtimestamp(self.excel_file.stat().st_mtime)
                if (self._file_modified_time and 
                    current_mtime == self._file_modified_time and
                    datetime.now() - self._last_load_time < self._cache_ttl):
                    # Cache is still valid, skip reload
                    return
                self._file_modified_time = current_mtime
            except (OSError, FileNotFoundError):
                # File doesn't exist or can't be accessed, reload anyway
                pass
        
        if self.excel_file.exists():
            try:
                wb = load_workbook(self.excel_file, read_only=True, data_only=True)
                if 'Customers' in wb.sheetnames:
                    sheet = wb['Customers']
                    self.customers = []
                    
                    # Read data starting from row 2 (row 1 is headers)
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        # Check if row has at least 6 columns and first column has a value
                        if len(row) >= 6 and row[0]:
                            try:
                                customer = {
                                    'name': str(row[0]).strip() if row[0] else '',
                                    'business': str(row[1]).strip() if row[1] else '',
                                    'address': str(row[2]).strip() if row[2] else '',
                                    'city': str(row[3]).strip() if row[3] else '',
                                    'state': str(row[4]).strip() if row[4] else '',
                                    'zip_code': str(row[5]).strip() if row[5] else ''
                                }
                                # Only add if name and business are not empty
                                if customer['name'] and customer['business']:
                                    self.customers.append(customer)
                            except (IndexError, TypeError) as e:
                                # Skip rows with invalid data
                                print(f"Warning: Skipping invalid customer row: {e}")
                                continue
                wb.close()
            except PermissionError as e:
                error_msg = (
                    f"ERROR CODE: CM001 - Customer Excel file is locked or open.\n"
                    f"File: {self.excel_file}\n"
                    f"TROUBLESHOOTING: Close Excel if file is open, check file permissions.\n"
                    f"Error: {e}"
                )
                print(f"WARNING: {error_msg}")
                # Keep existing customers if file is locked
                if not self.customers:
                    self.customers = []
            except Exception as e:
                error_msg = (
                    f"ERROR CODE: CM002 - Failed to load customers from Excel.\n"
                    f"File: {self.excel_file}\n"
                    f"TROUBLESHOOTING: Check if file exists, verify Excel format is valid.\n"
                    f"Error: {e}"
                )
                print(f"ERROR: {error_msg}")
                self.customers = []
        
        # Update cache timestamp
        self._last_load_time = datetime.now()
        if self.excel_file.exists():
            try:
                self._file_modified_time = datetime.fromtimestamp(self.excel_file.stat().st_mtime)
            except (OSError, FileNotFoundError):
                pass
        
        # If no customers exist, create Excel file with default customer
        if not self.customers:
            self._create_excel_with_default()
    
    def _create_excel_with_default(self):
        """Create Excel file with default customer (Josh Atwood)"""
        try:
            # Ensure directory exists
            self.excel_file.parent.mkdir(parents=True, exist_ok=True)
            
            # If file exists but is corrupted, try to recreate it
            if self.excel_file.exists():
                try:
                    # Try to open it to see if it's valid
                    test_wb = load_workbook(self.excel_file, read_only=True)
                    if 'Customers' not in test_wb.sheetnames:
                        # File exists but doesn't have Customers sheet - recreate it
                        test_wb.close()
                        self.excel_file.unlink()  # Delete corrupted file
                    else:
                        test_wb.close()
                        return  # File is valid, don't recreate
                except Exception:
                    # File is corrupted, delete and recreate
                    try:
                        self.excel_file.unlink()
                    except Exception:
                        pass
            
            # Create workbook
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Customers"
            
            # Add headers
            headers = ["Name", "Business", "Address", "City", "State", "Zip Code"]
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
            
            # Add default customer
            default_customer = ['Josh Atwood', 'Future Solutions Inc', '2616 Glenview Dr,', 'Elkhart', 'IN', '46514']
            for col_idx, value in enumerate(default_customer, start=1):
                sheet.cell(row=2, column=col_idx).value = value
            
            # Auto-adjust column widths
            for col in sheet.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[col_letter].width = adjusted_width
            
            wb.save(self.excel_file)
            wb.close()
            
            # Reload customers
            self._load_customers()
        except Exception as e:
            print(f"Error creating customer Excel file: {e}")
            import traceback
            traceback.print_exc()
    
    def _save_customers(self):
        """Save customers to Excel file (not used - users edit Excel directly)"""
        # This method is kept for compatibility but customers are edited directly in Excel
        pass
    
    def get_customers(self) -> List[Dict]:
        """Get list of all customers"""
        return self.customers.copy()
    
    def get_customer_names(self) -> List[str]:
        """Get list of customer display names (Name | Business)"""
        return [f"{c['name']} | {c['business']}" for c in self.customers]
    
    def get_customer_by_name(self, display_name: str) -> Optional[Dict]:
        """Get customer by display name (Name | Business)"""
        for customer in self.customers:
            if f"{customer['name']} | {customer['business']}" == display_name:
                return customer
        return None
    
    def format_customer_for_cell(self, customer: Dict) -> str:
        """
        Format customer information for Cell A3.
        
        Format: Each field on a new line
        Example:
        Josh Atwood
        Future Solutions Inc
        2616 Glenview Dr,
        Elkhart, IN 46514
        """
        return f"{customer['name']}\n{customer['business']}\n{customer['address']}\n{customer['city']}, {customer['state']} {customer['zip_code']}"
    
    def refresh_customers(self, force_reload=True):
        """
        Reload customers from Excel file (call this after manual edits).
        
        Args:
            force_reload: If True, bypass cache and reload from file (default: True)
        """
        self._load_customers(force_reload=force_reload)
    
    def add_customer(self, name: str, business: str, address: str, city: str, state: str, zip_code: str) -> bool:
        """
        Add a new customer to Excel file.
        
        Args:
            name: Customer name
            business: Business name
            address: Street address
            city: City
            state: State abbreviation
            zip_code: ZIP code
            
        Returns:
            True if added successfully, False if customer already exists
        """
        # Check if customer already exists (by name and business)
        for customer in self.customers:
            if customer['name'] == name and customer['business'] == business:
                return False
        
        try:
            # Load workbook
            wb = load_workbook(self.excel_file)
            if 'Customers' not in wb.sheetnames:
                wb.close()
                print("Error: 'Customers' sheet not found in Excel file")
                return False
            
            sheet = wb['Customers']
            
            # Find next empty row
            next_row = sheet.max_row + 1
            
            # Add customer data
            sheet.cell(row=next_row, column=1).value = name
            sheet.cell(row=next_row, column=2).value = business
            sheet.cell(row=next_row, column=3).value = address
            sheet.cell(row=next_row, column=4).value = city
            sheet.cell(row=next_row, column=5).value = state
            sheet.cell(row=next_row, column=6).value = zip_code
            
            wb.save(self.excel_file)
            wb.close()
            
            # Reload customers
            self._load_customers()
            return True
        except PermissionError as e:
            error_msg = (
                f"ERROR CODE: CM003 - Excel file is locked or open when adding customer.\n"
                f"File: {self.excel_file}\n"
                f"TROUBLESHOOTING: Close Excel if file is open, check file permissions.\n"
                f"Error: {e}"
            )
            print(f"ERROR: {error_msg}")
            return False
        except Exception as e:
            error_msg = (
                f"ERROR CODE: CM004 - Failed to add customer to Excel.\n"
                f"File: {self.excel_file}\n"
                f"Customer: {name} | {business}\n"
                f"TROUBLESHOOTING: Verify Excel file format, check file permissions.\n"
                f"Error: {e}"
            )
            print(f"ERROR: {error_msg}")
            return False
    
    def remove_customer(self, display_name: str) -> bool:
        """
        Remove a customer from Excel file by display name.
        
        Args:
            display_name: Display name (Name | Business)
            
        Returns:
            True if removed, False if not found
        """
        # Find customer to remove
        customer_to_remove = None
        for customer in self.customers:
            if f"{customer['name']} | {customer['business']}" == display_name:
                customer_to_remove = customer
                break
        
        if not customer_to_remove:
            return False
        
        try:
            # Load workbook
            wb = load_workbook(self.excel_file)
            if 'Customers' not in wb.sheetnames:
                wb.close()
                print("Error: 'Customers' sheet not found in Excel file")
                return False
            
            sheet = wb['Customers']
            
            # Find and delete row
            for row_idx in range(2, sheet.max_row + 1):
                cell1_value = sheet.cell(row=row_idx, column=1).value
                cell2_value = sheet.cell(row=row_idx, column=2).value
                
                # Handle None values
                name_match = (str(cell1_value).strip() if cell1_value is not None else '') == customer_to_remove['name']
                business_match = (str(cell2_value).strip() if cell2_value is not None else '') == customer_to_remove['business']
                
                if name_match and business_match:
                    sheet.delete_rows(row_idx)
                    break
            
            wb.save(self.excel_file)
            wb.close()
            
            # Reload customers
            self._load_customers()
            return True
        except PermissionError as e:
            error_msg = (
                f"ERROR CODE: CM005 - Excel file is locked or open when removing customer.\n"
                f"File: {self.excel_file}\n"
                f"Customer: {display_name}\n"
                f"TROUBLESHOOTING: Close Excel if file is open, check file permissions.\n"
                f"Error: {e}"
            )
            print(f"ERROR: {error_msg}")
            return False
        except Exception as e:
            error_msg = (
                f"ERROR CODE: CM006 - Failed to remove customer from Excel.\n"
                f"File: {self.excel_file}\n"
                f"Customer: {display_name}\n"
                f"TROUBLESHOOTING: Verify Excel file format, check file permissions.\n"
                f"Error: {e}"
            )
            print(f"ERROR: {error_msg}")
            return False
    
    def update_customer(self, old_display_name: str, name: str, business: str, address: str, city: str, state: str, zip_code: str) -> bool:
        """
        Update an existing customer's information in Excel file.
        
        Args:
            old_display_name: Current display name (Name | Business) of customer to update
            name: New customer name
            business: New business name
            address: New street address
            city: New city
            state: New state abbreviation
            zip_code: New ZIP code
            
        Returns:
            True if updated successfully, False if customer not found or update failed
        """
        # Find customer to update
        customer_to_update = None
        for customer in self.customers:
            if f"{customer['name']} | {customer['business']}" == old_display_name:
                customer_to_update = customer
                break
        
        if not customer_to_update:
            return False
        
        try:
            # Load workbook
            wb = load_workbook(self.excel_file)
            if 'Customers' not in wb.sheetnames:
                wb.close()
                print("Error: 'Customers' sheet not found in Excel file")
                return False
            
            sheet = wb['Customers']
            
            # Find and update row
            for row_idx in range(2, sheet.max_row + 1):
                cell1_value = sheet.cell(row=row_idx, column=1).value
                cell2_value = sheet.cell(row=row_idx, column=2).value
                
                # Handle None values
                name_match = (str(cell1_value).strip() if cell1_value is not None else '') == customer_to_update['name']
                business_match = (str(cell2_value).strip() if cell2_value is not None else '') == customer_to_update['business']
                
                if name_match and business_match:
                    # Update customer data
                    sheet.cell(row=row_idx, column=1).value = name
                    sheet.cell(row=row_idx, column=2).value = business
                    sheet.cell(row=row_idx, column=3).value = address
                    sheet.cell(row=row_idx, column=4).value = city
                    sheet.cell(row=row_idx, column=5).value = state
                    sheet.cell(row=row_idx, column=6).value = zip_code
                    break
            
            wb.save(self.excel_file)
            wb.close()
            
            # Reload customers
            self._load_customers()
            return True
        except PermissionError as e:
            error_msg = (
                f"ERROR CODE: CM007 - Excel file is locked or open when updating customer.\n"
                f"File: {self.excel_file}\n"
                f"TROUBLESHOOTING: Close Excel if file is open, check file permissions.\n"
                f"Error: {e}"
            )
            print(f"ERROR: {error_msg}")
            return False
        except Exception as e:
            error_msg = (
                f"ERROR CODE: CM008 - Failed to update customer in Excel.\n"
                f"File: {self.excel_file}\n"
                f"Customer: {old_display_name}\n"
                f"TROUBLESHOOTING: Verify Excel file format, check file permissions.\n"
                f"Error: {e}"
            )
            print(f"ERROR: {error_msg}")
            return False
    
    def open_excel_file(self):
        """Open the Excel file for manual editing"""
        import subprocess
        import platform
        
        try:
            if platform.system() == 'Windows':
                subprocess.run(['start', '', str(self.excel_file.absolute())], 
                             shell=True, check=False)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', str(self.excel_file.absolute())], 
                             check=False)
            else:  # Linux
                subprocess.run(['xdg-open', str(self.excel_file.absolute())], 
                             check=False)
        except Exception as e:
            print(f"Error opening Excel file: {e}")

