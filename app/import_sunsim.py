#!/usr/bin/env python3
"""
Sun Simulator Import Tool for Packout Operations

This script imports sun simulator export files and updates the current
pallet Excel workbook so packout operators can scan barcodes and automatically
populate electrical values.

OFFLINE ONLY - No network connections required.
"""

import sys
import argparse
import logging
import re
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Tuple
import pandas as pd
from openpyxl import load_workbook

# Try to import python-dateutil, fall back to pandas if not available
try:
    from dateutil import parser as date_parser
    DATEUTIL_AVAILABLE = True
except ImportError:
    DATEUTIL_AVAILABLE = False
    date_parser = None

# Configuration
SCRIPT_DIR = Path(__file__).parent.absolute()
CSV_INPUT_DIR = SCRIPT_DIR / "CSV_INPUT"
EXCEL_DIR = SCRIPT_DIR / "EXCEL"
BACKUPS_DIR = EXCEL_DIR / "backups"
LOGS_DIR = SCRIPT_DIR / "LOGS"
ARCHIVE_DIR = SCRIPT_DIR / "ARCHIVE" / "processed_files"

# Required columns mapping
REQUIRED_FIELDS = {
    'SerialNo': ['SerialNo', 'serialno', 'Serial Number', 'serial number'],
    'Pm': ['Pm', 'pm', 'Pm(W)', 'Power'],
    'Isc': ['Isc', 'isc', 'Isc(A)', 'Short Circuit Current'],
    'Voc': ['Voc(V)', 'Voc', 'voc', 'voc(v)', 'Open Circuit Voltage'],
    'Ipm': ['Ipm', 'ipm', 'Ipm(A)', 'Current at Max Power'],
    'Vpm': ['Vpm(V)', 'Vpm', 'vpm', 'vpm(v)', 'Voltage at Max Power']
}

OPTIONAL_FIELDS = {
    'Date': ['Date', 'date', 'DATE'],
    'TTime': ['TTime', 'ttime', 'T Time', 'Time']
}

# Excel column mapping (DATA sheet)
EXCEL_COLUMNS = {
    'SerialNo': 'B',
    'Date': 'C',
    'TTime': 'D',
    'Pm': 'H',
    'Isc': 'I',
    'Voc': 'J',
    'Ipm': 'K',
    'Vpm': 'L'
}

# Value ranges for validation (warn but don't block)
VALUE_RANGES = {
    'Pm': (0, 1000),   # Watts
    'Voc': (0, 100),   # Volts
    'Vpm': (0, 100),   # Volts
    'Isc': (0, 50),    # Amps
    'Ipm': (0, 50),    # Amps
}

# Panel type expected power ranges (from specification table)
# Format: Panel Type (Watts): (Minimum Output, Maximum Output)
PANEL_TYPE_RANGES = {
    200: (195, 206),
    220: (214, 227),
    325: (316, 335),
    335: (325, 345.05),
    380: (369, 391.4),
    385: (375, 396.55),
    390: (379, 401.7),
    395: (384, 406.85),
    450: (439, 463.5),
}


def setup_logging() -> logging.Logger:
    """Set up logging to LOGS/import_log.txt"""
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOGS_DIR / "import_log.txt"
    
    logger = logging.getLogger('sunsim_import')
    logger.setLevel(logging.INFO)
    
    # File handler
    file_handler = logging.FileHandler(log_file, mode='a', encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    
    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    
    # Formatter
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    return logger


def find_newest_input_file() -> Optional[Path]:
    """
    Find the newest file in CSV_INPUT/ directory.
    Supports both .csv and .xlsx files.
    """
    if not CSV_INPUT_DIR.exists():
        return None
    
    files = []
    for ext in ['.csv', '.xlsx']:
        files.extend(list(CSV_INPUT_DIR.glob(f'*{ext}')))
    
    if not files:
        return None
    
    # Return most recently modified file
    return max(files, key=lambda p: p.stat().st_mtime)


def normalize_field_name(field_name: str) -> str:
    """Normalize field name by stripping whitespace and removing Excel wrappers"""
    if not field_name:
        return ""
    
    # Remove Excel-style ="VALUE" wrappers
    field_name = str(field_name).strip()
    if field_name.startswith('="') and field_name.endswith('"'):
        field_name = field_name[2:-1]
    
    return field_name.strip()


def parse_barcode_panel_type(serial_no: str) -> Optional[int]:
    """
    Parse barcode to extract panel type.
    Format: CRSYYFBPP####
    - CRS: Company prefix (fixed) - positions 0-2
    - YY: Year code (2 digits) - positions 3-4
    - F: Frame type (W/B) - position 5
    - B: Backsheet type (T/W/B) - position 6
    - PP: Panel type (36, 40, 60, 72, or 144) - positions 7-8 or 7-9
    - ####: Sequential number (4-6 digits) - positions 9+ or 10+
    
    Example: CRS25BW1440001
    - CRS = "CRS" (0-2)
    - 25 = "25" (3-4)
    - B = "B" (5) - Black frame
    - W = "W" (6) - White backsheet
    - 144 = "144" (7-9) - Panel type
    - 0001 = "0001" (10+) - Sequence
    
    Returns panel type as integer, or None if parsing fails.
    """
    if not serial_no or len(serial_no) < 9:
        return None
    
    # Check for CRS prefix
    if not serial_no.startswith('CRS'):
        return None
    
    # Need at least 9 characters: CRS + YY + F + B + PP (min 2 digits)
    if len(serial_no) < 9:
        return None
    
    try:
        # Panel type starts at position 7 (after CRSYYFB)
        # First try 3-digit panel type (144)
        if len(serial_no) >= 10:
            panel_type_str = serial_no[7:10]
            try:
                panel_type = int(panel_type_str)
                if panel_type == 144:
                    return panel_type
            except ValueError:
                pass
        
        # Try 2-digit panel type (36, 40, 60, 72)
        if len(serial_no) >= 9:
            panel_type_str = serial_no[7:9]
            try:
                panel_type = int(panel_type_str)
                # Check if it's a valid 2-digit panel type
                if panel_type in [36, 40, 60, 72]:
                    return panel_type
                # Also check if it matches any known panel type from ranges
                if panel_type in PANEL_TYPE_RANGES:
                    return panel_type
            except ValueError:
                pass
                
    except (ValueError, IndexError):
        pass
    
    return None


def normalize_numeric(value) -> Optional[float]:
    """Convert value to float, handling various formats"""
    if value is None or value == '':
        return None
    
    # Remove Excel wrappers
    if isinstance(value, str):
        value = normalize_field_name(value)
        if value == '':
            return None
    
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_csv_file(file_path: Path, logger: logging.Logger) -> List[Dict]:
    """Parse CSV file and extract required fields"""
    records = []
    
    try:
        # Try different encodings
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                df = pd.read_csv(file_path, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            logger.error(f"Could not decode CSV file {file_path} with any encoding")
            return records
        
        # Normalize column names
        df.columns = [normalize_field_name(col) for col in df.columns]
        
        # Find required columns
        field_mapping = {}
        for target_field, possible_names in {**REQUIRED_FIELDS, **OPTIONAL_FIELDS}.items():
            for col in df.columns:
                if col in possible_names:
                    field_mapping[target_field] = col
                    break
        
        # Check if DataFrame is empty
        if df.empty:
            logger.warning(f"CSV file {file_path.name} contains no data rows")
            return records
        
        # Extract records - use itertuples for better performance (3-5x faster)
        for row in df.itertuples(index=False):
            record = {}
            # Get column names, handling spaces in column names
            col_names = [col.replace(' ', '_') if ' ' in col else col for col in df.columns]
            
            for target_field in ['SerialNo', 'Date', 'TTime', 'Pm', 'Isc', 'Voc', 'Ipm', 'Vpm']:
                source_col = field_mapping.get(target_field)
                if source_col and source_col in df.columns:
                    try:
                        # Get value from named tuple
                        safe_col_name = source_col.replace(' ', '_') if ' ' in source_col else source_col
                        value = getattr(row, safe_col_name, None)
                        if value is None:
                            # Try original column name
                            col_idx = list(df.columns).index(source_col)
                            value = row[col_idx] if col_idx < len(row) else None
                        if pd.isna(value) if hasattr(pd, 'isna') else (value is None or (isinstance(value, float) and str(value) == 'nan')):
                            value = None
                    except (AttributeError, IndexError):
                        value = None
                    
                    if target_field == 'SerialNo':
                        record[target_field] = normalize_field_name(str(value)) if value is not None else ""
                    elif target_field in ['Date', 'TTime']:
                        record[target_field] = normalize_field_name(str(value)) if value is not None else None
                    else:
                        record[target_field] = normalize_numeric(value)
                else:
                    record[target_field] = None if target_field in ['Date', 'TTime'] else (None if target_field != 'SerialNo' else "")
            
            records.append(record)
        
        logger.info(f"Parsed {len(records)} records from CSV file {file_path.name}")
        
    except Exception as e:
        logger.error(f"Error parsing CSV file {file_path}: {e}")
    
    return records


def parse_xlsx_file(file_path: Path, logger: logging.Logger) -> List[Dict]:
    """Parse XLSX file and extract required fields"""
    records = []
    
    try:
        # Read first sheet
        df = pd.read_excel(file_path, sheet_name=0)
        
        # Normalize column names
        df.columns = [normalize_field_name(col) for col in df.columns]
        
        # Find required columns
        field_mapping = {}
        for target_field, possible_names in {**REQUIRED_FIELDS, **OPTIONAL_FIELDS}.items():
            matches = [col for col in df.columns if col in possible_names]
            if matches:
                # Use first match, but warn if multiple matches found
                if len(matches) > 1:
                    logger.warning(f"Multiple columns match '{target_field}': {matches}. Using '{matches[0]}'")
                field_mapping[target_field] = matches[0]
        
        # Check if DataFrame is empty
        if df.empty:
            logger.warning(f"XLSX file {file_path.name} contains no data rows")
            return records
        
        # Extract records - use itertuples for better performance (3-5x faster)
        for row in df.itertuples(index=False):
            record = {}
            # Get column names, handling spaces in column names
            col_names = [col.replace(' ', '_') if ' ' in col else col for col in df.columns]
            
            for target_field in ['SerialNo', 'Date', 'TTime', 'Pm', 'Isc', 'Voc', 'Ipm', 'Vpm']:
                source_col = field_mapping.get(target_field)
                if source_col and source_col in df.columns:
                    try:
                        # Get value from named tuple
                        safe_col_name = source_col.replace(' ', '_') if ' ' in source_col else source_col
                        value = getattr(row, safe_col_name, None)
                        if value is None:
                            # Try original column name
                            col_idx = list(df.columns).index(source_col)
                            value = row[col_idx] if col_idx < len(row) else None
                        if pd.isna(value) if hasattr(pd, 'isna') else (value is None or (isinstance(value, float) and str(value) == 'nan')):
                            value = None
                    except (AttributeError, IndexError):
                        value = None
                    
                    if target_field == 'SerialNo':
                        record[target_field] = normalize_field_name(str(value)) if value is not None else ""
                    elif target_field in ['Date', 'TTime']:
                        record[target_field] = normalize_field_name(str(value)) if value is not None else None
                    else:
                        record[target_field] = normalize_numeric(value)
                else:
                    record[target_field] = None if target_field in ['Date', 'TTime'] else (None if target_field != 'SerialNo' else "")
            
            records.append(record)
        
        logger.info(f"Parsed {len(records)} records from XLSX file {file_path.name}")
        
    except Exception as e:
        logger.error(f"Error parsing XLSX file {file_path}: {e}")
    
    return records


def validate_serialno_format(serial_no: str, logger: logging.Logger, row_num: Optional[int] = None) -> List[str]:
    """
    Validate SerialNo format and detect suspicious patterns.
    Returns list of warnings (empty if no issues).
    """
    warnings = []
    row_info = f" (row {row_num})" if row_num else ""
    
    if not serial_no:
        return warnings
    
    # Check length (reasonable range for CRS format: 13-17 characters)
    # Format: CRSYYFBPP#### (min 13: CRS + YY + F + B + PP + 4 digits)
    #         CRSYYFBPP###### (max ~17: CRS + YY + F + B + 144 + 6 digits)
    if len(serial_no) < 10:
        warnings.append(f"SerialNo '{serial_no}' is unusually short (length {len(serial_no)}){row_info}")
        logger.warning(f"SerialNo '{serial_no}' is unusually short{row_info}")
    elif len(serial_no) > 20:
        warnings.append(f"SerialNo '{serial_no}' is unusually long (length {len(serial_no)}){row_info}")
        logger.warning(f"SerialNo '{serial_no}' is unusually long{row_info}")
    
    # Check for suspicious patterns
    # All zeros
    if serial_no.replace('0', '').strip() == '':
        warnings.append(f"SerialNo '{serial_no}' contains only zeros{row_info}")
        logger.warning(f"Suspicious SerialNo pattern: all zeros '{serial_no}'{row_info}")
    
    # All same character (except if it's a valid pattern)
    if len(set(serial_no)) == 1 and len(serial_no) > 3:
        warnings.append(f"SerialNo '{serial_no}' contains only repeated characters{row_info}")
        logger.warning(f"Suspicious SerialNo pattern: all same character '{serial_no}'{row_info}")
    
    # Many repeated characters (e.g., "CRS25BW1441111")
    if len(serial_no) >= 4:
        # Check if last 4+ characters are all the same
        last_chars = serial_no[-4:]
        if len(set(last_chars)) == 1:
            warnings.append(f"SerialNo '{serial_no}' has suspicious repeated pattern at end{row_info}")
            logger.warning(f"Suspicious SerialNo pattern: repeated characters at end '{serial_no}'{row_info}")
    
    # Check for CRS format but invalid structure
    if serial_no.startswith('CRS'):
        if len(serial_no) < 13:
            warnings.append(f"CRS format SerialNo '{serial_no}' appears incomplete (expected ~13-17 chars){row_info}")
            logger.warning(f"CRS format SerialNo '{serial_no}' appears incomplete{row_info}")
    
    return warnings


def validate_record(record: Dict, logger: logging.Logger, row_num: Optional[int] = None, verbose: bool = False, exclude_out_of_range: bool = False) -> Tuple[bool, List[str]]:
    """
    Validate that record has all required fields and check value ranges.
    If exclude_out_of_range is True, will exclude records where Pm is outside panel type range.
    Returns (is_valid, list_of_warnings)
    """
    warnings = []
    row_info = f" (row {row_num})" if row_num else ""
    
    # Check SerialNo - must be non-empty string
    serial_no = record.get('SerialNo')
    if not serial_no:
        if verbose:
            logger.warning(f"Missing SerialNo{row_info}")
        return False, ["Missing SerialNo"]
    # Convert to string if it's not already (handles numeric SerialNos)
    if not isinstance(serial_no, str):
        serial_no = str(serial_no)
    if serial_no.strip() == '':
        if verbose:
            logger.warning(f"Empty SerialNo{row_info}")
        return False, ["Empty SerialNo"]
    
    # Validate SerialNo format and check for suspicious patterns
    format_warnings = validate_serialno_format(serial_no, logger, row_num)
    warnings.extend(format_warnings)
    
    # Check all numeric fields
    required_numeric = ['Pm', 'Isc', 'Voc', 'Ipm', 'Vpm']
    for field in required_numeric:
        value = record.get(field)
        if value is None:
            if verbose:
                logger.warning(f"Missing {field} for SerialNo {serial_no}{row_info}")
            return False, [f"Missing {field}"]
        
        # Check value ranges
        if field in VALUE_RANGES:
            min_val, max_val = VALUE_RANGES[field]
            if value < min_val or value > max_val:
                warning = f"{field} value {value} is outside expected range ({min_val}-{max_val}) for SerialNo {serial_no}{row_info}"
                warnings.append(warning)
                logger.warning(warning)
    
    # Panel type validation - check if Pm is within expected range for panel type
    panel_type = parse_barcode_panel_type(serial_no)
    if panel_type and panel_type in PANEL_TYPE_RANGES:
        pm_value = record.get('Pm')
        if pm_value is not None:
            min_power, max_power = PANEL_TYPE_RANGES[panel_type]
            if pm_value < min_power or pm_value > max_power:
                warning_msg = f"Panel type {panel_type} SerialNo {serial_no} has Pm={pm_value}W outside expected range ({min_power}-{max_power}W){row_info}"
                warnings.append(warning_msg)
                logger.warning(warning_msg)
                
                # If exclude_out_of_range is True, exclude this record
                if exclude_out_of_range:
                    logger.error(f"EXCLUDING SerialNo {serial_no} - power {pm_value}W outside panel type {panel_type} range ({min_power}-{max_power}W){row_info}")
                    return False, [f"Power {pm_value}W outside panel type {panel_type} range ({min_power}-{max_power}W)"]
    elif panel_type is None and verbose:
        # Couldn't parse panel type from barcode
        logger.debug(f"Could not parse panel type from SerialNo {serial_no}{row_info}")
    
    return True, warnings


def parse_date_time(date_str: Optional[str], time_str: Optional[str], logger: logging.Logger, row_num: Optional[int] = None) -> Tuple[float, float, List[str]]:
    """
    Parse date and time strings to timestamps using flexible parsing.
    Returns (date_timestamp, time_timestamp, list_of_warnings)
    """
    warnings = []
    row_info = f" (row {row_num})" if row_num else ""
    date_val = 0
    time_val = 0
    
    # Parse date
    if date_str:
        date_str = str(date_str).strip()
        if date_str:
            try:
                if DATEUTIL_AVAILABLE:
                    # Use python-dateutil for flexible parsing
                    parsed_date = date_parser.parse(date_str, fuzzy=True, default=datetime(2000, 1, 1))
                else:
                    # Fall back to pandas
                    parsed_date = pd.to_datetime(date_str)
                
                date_val = parsed_date.timestamp()
                
                # Check for future dates (likely errors)
                now = datetime.now().timestamp()
                if date_val > now:
                    warning = f"Future date detected: {date_str} (row {row_num})" if row_num else f"Future date detected: {date_str}"
                    warnings.append(warning)
                    logger.warning(warning)
                    
            except (ValueError, TypeError, pd.errors.ParserError) as e:
                warning = f"Could not parse date '{date_str}'{row_info}: {e}"
                warnings.append(warning)
                logger.debug(warning)
    
    # Parse time
    if time_str:
        time_str = str(time_str).strip()
        if time_str:
            try:
                if DATEUTIL_AVAILABLE:
                    # Use python-dateutil for flexible parsing
                    parsed_time = date_parser.parse(time_str, fuzzy=True, default=datetime(2000, 1, 1))
                else:
                    # Fall back to pandas
                    parsed_time = pd.to_datetime(time_str)
                
                time_val = parsed_time.timestamp()
                
            except (ValueError, TypeError, pd.errors.ParserError) as e:
                warning = f"Could not parse time '{time_str}'{row_info}: {e}"
                warnings.append(warning)
                logger.debug(warning)
    
    return date_val, time_val, warnings


def deduplicate_records(records: List[Dict], file_mtime: float, logger: logging.Logger) -> Dict[str, Dict]:
    """
    Apply latest-run-wins logic to deduplicate records by SerialNo.
    Returns dictionary mapping SerialNo to best record.
    """
    serial_map = {}
    all_warnings = []
    
    for record in records:
        serial_no = record.get('SerialNo', '').strip()
        if not serial_no:
            continue
        
        # Get date and time for sorting
        date_str = record.get('Date') or ''
        ttime_str = record.get('TTime') or ''
        
        # Parse date and time with improved parsing
        date_val, time_val, warnings = parse_date_time(date_str, ttime_str, logger)
        all_warnings.extend(warnings)
        
        # Create sort key: (file_mtime, date, time)
        # Higher values = more recent
        sort_key = (file_mtime, date_val, time_val)
        
        # Keep the record with the highest sort key
        if serial_no not in serial_map:
            serial_map[serial_no] = (sort_key, record)
        else:
            existing_key, existing_record = serial_map[serial_no]
            if sort_key > existing_key:
                serial_map[serial_no] = (sort_key, record)
    
    # Log summary of date/time parsing issues if any
    if all_warnings and logger:
        logger.debug(f"Date/time parsing: {len(all_warnings)} warnings during deduplication")
    
    # Return just the records
    return {serial: record for serial, (_, record) in serial_map.items()}


def find_pallet_workbook(override_path: Optional[str] = None, logger: Optional[logging.Logger] = None) -> Optional[Path]:
    """
    Find the target pallet workbook.
    Priority:
    1. Override path (--xlsx argument)
    2. EXCEL/CURRENT.xlsx
    3. Most recently modified BUILD YYYY Q-X.xlsx file
    """
    if override_path:
        path = Path(override_path)
        if path.exists():
            if logger:
                logger.info(f"Using override workbook: {path}")
            return path
        else:
            if logger:
                logger.error(f"Override workbook not found: {override_path}")
            return None
    
    # Check for CURRENT.xlsx
    current_file = EXCEL_DIR / "CURRENT.xlsx"
    if current_file.exists():
        if logger:
            logger.info(f"Using CURRENT.xlsx: {current_file}")
        return current_file
    
    # Find BUILD files
    build_files = []
    for file in EXCEL_DIR.glob("*.xlsx"):
        # Skip temp files
        if file.name.startswith("~$"):
            continue
        
        # Check for BUILD pattern
        name = file.name
        if "BUILD" in name.upper():
            # Check for year (4 digits)
            if re.search(r'\d{4}', name):
                # Check for quarter pattern
                if re.search(r'Q\s*-?\s*\d', name, re.IGNORECASE):
                    build_files.append(file)
    
    if not build_files:
        if logger:
            logger.error("No pallet workbook found in EXCEL/ directory")
        return None
    
    # Return most recently modified
    latest = max(build_files, key=lambda p: p.stat().st_mtime)
    if logger:
        logger.info(f"Found pallet workbook: {latest.name}")
    return latest


def create_backup(workbook_path: Path, logger: logging.Logger) -> Optional[Path]:
    """Create timestamped backup of workbook"""
    BACKUPS_DIR.mkdir(parents=True, exist_ok=True)
    
    # Create unique backup filename with timestamp and counter if needed
    base_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{workbook_path.stem}_backup_{base_timestamp}.xlsx"
    backup_path = BACKUPS_DIR / backup_name
    
    # Handle filename collision (if same second)
    counter = 1
    while backup_path.exists():
        backup_name = f"{workbook_path.stem}_backup_{base_timestamp}_{counter}.xlsx"
        backup_path = BACKUPS_DIR / backup_name
        counter += 1
    
    try:
        # Copy workbook
        wb = load_workbook(workbook_path)
        wb.save(backup_path)
        logger.info(f"Created backup: {backup_path}")
        return backup_path
    except PermissionError as e:
        logger.error(f"Failed to create backup - file may be open in Excel: {e}")
        return None
    except Exception as e:
        logger.error(f"Failed to create backup: {e}")
        return None


def validate_workbook_structure(workbook_path: Path, logger: logging.Logger) -> Tuple[bool, List[str]]:
    """
    Validate workbook structure before processing.
    Returns (is_valid, list_of_warnings)
    """
    warnings = []
    
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        
        # Check for DATA sheet
        if 'DATA' not in wb.sheetnames:
            logger.error("DATA sheet not found in workbook")
            return False, ["DATA sheet not found"]
        
        data_sheet = wb['DATA']
        
        # Check for required columns by examining first few rows
        required_cols = ['B', 'H', 'I', 'J', 'K', 'L']
        missing_cols = []
        
        # Check if we can read from required columns (they exist)
        try:
            # Try to read from each required column
            for col in required_cols:
                cell = data_sheet[f'{col}2']  # Check row 2 (assuming row 1 is header)
                # If we can access it, column exists
        except Exception:
            # If we can't access, might be an issue but continue
            pass
        
        # Check if PALLET SHEET exists (should not be modified)
        if 'PALLET SHEET' not in wb.sheetnames:
            warnings.append("PALLET SHEET not found - VLOOKUP formulas may not work")
            logger.warning("PALLET SHEET not found in workbook")
        
        wb.close()
        
        if warnings:
            logger.warning(f"Workbook structure warnings: {', '.join(warnings)}")
        
        return True, warnings
        
    except Exception as e:
        logger.error(f"Error validating workbook structure: {e}")
        return False, [f"Error reading workbook: {e}"]


def update_excel_workbook(workbook_path: Path, records: Dict[str, Dict], logger: logging.Logger, dry_run: bool = False, show_progress: bool = True) -> Tuple[int, int, List[Dict]]:
    """
    Update Excel workbook with new records.
    Returns (updated_count, added_count, change_details)
    change_details is a list of dicts with change information for reporting
    """
    change_details = []
    try:
        wb = load_workbook(workbook_path)
        
        # Get DATA sheet
        if 'DATA' not in wb.sheetnames:
            logger.error("DATA sheet not found in workbook")
            return 0, 0
        
        data_sheet = wb['DATA']
        
        # Build SerialNo to row mapping (column B)
        serial_to_row = {}
        for row_idx, row in enumerate(data_sheet.iter_rows(min_row=2, max_col=2, values_only=False), start=2):
            # Check if row has at least 2 columns (A and B)
            if len(row) < 2:
                continue
            cell = row[1]  # Column B
            if cell.value:
                serial = normalize_field_name(str(cell.value))
                if serial:
                    # If SerialNo already exists, log warning but keep last occurrence
                    if serial in serial_to_row:
                        logger.warning(f"Duplicate SerialNo '{serial}' found in Excel at row {row_idx} (previous at row {serial_to_row[serial]})")
                    serial_to_row[serial] = row_idx
        
        updated_count = 0
        added_count = 0
        total_records = len(records)
        
        # Process each record
        for idx, (serial_no, record) in enumerate(records.items(), 1):
            # Show progress
            if show_progress and (idx % 10 == 0 or idx == total_records):
                percent = (idx * 100) // total_records
                print(f"Processing {idx}/{total_records} ({percent}%)...", end='\r')
            
            if serial_no in serial_to_row:
                # Update existing row
                row_idx = serial_to_row[serial_no]
                
                # Get before values for report
                before_values = {
                    'Pm': data_sheet[f'{EXCEL_COLUMNS["Pm"]}{row_idx}'].value,
                    'Isc': data_sheet[f'{EXCEL_COLUMNS["Isc"]}{row_idx}'].value,
                    'Voc': data_sheet[f'{EXCEL_COLUMNS["Voc"]}{row_idx}'].value,
                    'Ipm': data_sheet[f'{EXCEL_COLUMNS["Ipm"]}{row_idx}'].value,
                    'Vpm': data_sheet[f'{EXCEL_COLUMNS["Vpm"]}{row_idx}'].value,
                    'Date': data_sheet[f'{EXCEL_COLUMNS["Date"]}{row_idx}'].value,
                    'TTime': data_sheet[f'{EXCEL_COLUMNS["TTime"]}{row_idx}'].value,
                }
                
                # Update columns H-L (Pm, Isc, Voc, Ipm, Vpm)
                data_sheet[f'{EXCEL_COLUMNS["Pm"]}{row_idx}'] = record.get('Pm')
                data_sheet[f'{EXCEL_COLUMNS["Isc"]}{row_idx}'] = record.get('Isc')
                data_sheet[f'{EXCEL_COLUMNS["Voc"]}{row_idx}'] = record.get('Voc')
                data_sheet[f'{EXCEL_COLUMNS["Ipm"]}{row_idx}'] = record.get('Ipm')
                data_sheet[f'{EXCEL_COLUMNS["Vpm"]}{row_idx}'] = record.get('Vpm')
                
                # Optionally update Date and TTime
                if record.get('Date'):
                    data_sheet[f'{EXCEL_COLUMNS["Date"]}{row_idx}'] = record.get('Date')
                if record.get('TTime'):
                    data_sheet[f'{EXCEL_COLUMNS["TTime"]}{row_idx}'] = record.get('TTime')
                
                # Record change details for report
                change_details.append({
                    'SerialNo': serial_no,
                    'Action': 'Updated',
                    'Row': row_idx,
                    'Before_Pm': before_values['Pm'],
                    'After_Pm': record.get('Pm'),
                    'Before_Isc': before_values['Isc'],
                    'After_Isc': record.get('Isc'),
                    'Before_Voc': before_values['Voc'],
                    'After_Voc': record.get('Voc'),
                    'Before_Ipm': before_values['Ipm'],
                    'After_Ipm': record.get('Ipm'),
                    'Before_Vpm': before_values['Vpm'],
                    'After_Vpm': record.get('Vpm'),
                    'Date': record.get('Date'),
                    'TTime': record.get('TTime'),
                })
                
                updated_count += 1
            else:
                # Append new row
                next_row = data_sheet.max_row + 1
                
                data_sheet[f'{EXCEL_COLUMNS["SerialNo"]}{next_row}'] = serial_no
                if record.get('Date'):
                    data_sheet[f'{EXCEL_COLUMNS["Date"]}{next_row}'] = record.get('Date')
                if record.get('TTime'):
                    data_sheet[f'{EXCEL_COLUMNS["TTime"]}{next_row}'] = record.get('TTime')
                data_sheet[f'{EXCEL_COLUMNS["Pm"]}{next_row}'] = record.get('Pm')
                data_sheet[f'{EXCEL_COLUMNS["Isc"]}{next_row}'] = record.get('Isc')
                data_sheet[f'{EXCEL_COLUMNS["Voc"]}{next_row}'] = record.get('Voc')
                data_sheet[f'{EXCEL_COLUMNS["Ipm"]}{next_row}'] = record.get('Ipm')
                data_sheet[f'{EXCEL_COLUMNS["Vpm"]}{next_row}'] = record.get('Vpm')
                
                # Record change details for report
                change_details.append({
                    'SerialNo': serial_no,
                    'Action': 'Added',
                    'Row': next_row,
                    'Before_Pm': None,
                    'After_Pm': record.get('Pm'),
                    'Before_Isc': None,
                    'After_Isc': record.get('Isc'),
                    'Before_Voc': None,
                    'After_Voc': record.get('Voc'),
                    'Before_Ipm': None,
                    'After_Ipm': record.get('Ipm'),
                    'Before_Vpm': None,
                    'After_Vpm': record.get('Vpm'),
                    'Date': record.get('Date'),
                    'TTime': record.get('TTime'),
                })
                
                added_count += 1
        
        if show_progress:
            print()  # New line after progress
        
        # Save workbook (unless dry-run)
        if not dry_run:
            try:
                wb.save(workbook_path)
                logger.info(f"Updated workbook: {updated_count} updated, {added_count} added")
            except PermissionError as e:
                logger.error(f"Cannot save workbook - file may be open in Excel: {e}")
                raise
            except Exception as e:
                logger.error(f"Error saving workbook: {e}")
                raise
        else:
            logger.info(f"[DRY-RUN] Would update workbook: {updated_count} updated, {added_count} added")
        
        return updated_count, added_count, change_details
        
    except PermissionError:
        # Re-raise permission errors so caller can handle them
        raise
    except Exception as e:
        logger.error(f"Error updating workbook: {e}")
        return 0, 0, []


def generate_summary_report(change_details: List[Dict], skipped_records: List[Dict], logger: logging.Logger, input_filename: str, workbook_name: str) -> Optional[Path]:
    """
    Generate CSV summary report of all changes made during import.
    Returns path to report file, or None if generation fails.
    """
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"import_report_{timestamp}.csv"
        report_path = LOGS_DIR / report_filename
        
        # Prepare report data
        report_rows = []
        
        # Add header
        report_rows.append({
            'Report_Generated': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'Input_File': input_filename,
            'Workbook': workbook_name,
            'Type': 'Header',
            'SerialNo': '',
            'Action': '',
            'Row': '',
            'Before_Pm': '',
            'After_Pm': '',
            'Before_Isc': '',
            'After_Isc': '',
            'Before_Voc': '',
            'After_Voc': '',
            'Before_Ipm': '',
            'After_Ipm': '',
            'Before_Vpm': '',
            'After_Vpm': '',
            'Date': '',
            'TTime': '',
            'Notes': '',
        })
        
        # Add change details
        for change in change_details:
            report_rows.append(change)
        
        # Add skipped records summary
        if skipped_records:
            report_rows.append({
                'SerialNo': f'SUMMARY: {len(skipped_records)} skipped records',
                'Action': 'Skipped',
                'Notes': 'See log file for details',
            })
            for skipped in skipped_records[:20]:  # First 20 skipped
                report_rows.append({
                    'SerialNo': skipped.get('SerialNo', 'N/A'),
                    'Action': 'Skipped',
                    'Notes': skipped.get('reason', 'Validation failed'),
                })
        
        # Write CSV
        import csv
        if report_rows:
            with open(report_path, 'w', newline='', encoding='utf-8') as f:
                if report_rows:
                    fieldnames = list(report_rows[0].keys())
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    for row in report_rows:
                        writer.writerow(row)
        
        logger.info(f"Generated summary report: {report_path}")
        return report_path
        
    except Exception as e:
        logger.error(f"Failed to generate summary report: {e}")
        return None


def archive_file(file_path: Path, logger: logging.Logger) -> bool:
    """Move processed file to archive"""
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    
    # Handle duplicate filenames by adding timestamp and counter
    archive_path = ARCHIVE_DIR / file_path.name
    if archive_path.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = file_path.stem
        suffix = file_path.suffix
        counter = 1
        archive_path = ARCHIVE_DIR / f"{stem}_{timestamp}{suffix}"
        while archive_path.exists():
            archive_path = ARCHIVE_DIR / f"{stem}_{timestamp}_{counter}{suffix}"
            counter += 1
    
    try:
        file_path.rename(archive_path)
        logger.info(f"Archived file to: {archive_path}")
        return True
    except Exception as e:
        logger.error(f"Failed to archive file: {e}")
        return False


def main():
    """Main execution function"""
    parser = argparse.ArgumentParser(description='Import sun simulator data into pallet workbook')
    parser.add_argument('--xlsx', type=str, help='Override pallet workbook path')
    parser.add_argument('--dry-run', action='store_true', help='Preview changes without saving')
    parser.add_argument('--verbose', '-v', action='store_true', help='Show detailed processing information')
    parser.add_argument('--exclude-out-of-range', action='store_true', help='Exclude records where power is outside panel type range')
    args = parser.parse_args()
    
    # Setup logging
    logger = setup_logging()
    logger.info("=" * 60)
    logger.info("Sun Simulator Import Tool - Starting")
    logger.info("=" * 60)
    
    # Find input file
    input_file = find_newest_input_file()
    if not input_file:
        logger.error("No input files found in CSV_INPUT/ directory")
        print("Error: No input files found in CSV_INPUT/ directory")
        return 1
    
    logger.info(f"Processing input file: {input_file.name}")
    
    # Parse input file
    if input_file.suffix.lower() == '.csv':
        records = parse_csv_file(input_file, logger)
    else:
        records = parse_xlsx_file(input_file, logger)
    
    if not records:
        logger.error("No records parsed from input file")
        print("Error: No records parsed from input file")
        return 1
    
    # Validate and filter records
    valid_records = []
    skipped_records = []  # Track skipped records with reasons for report
    skipped_count = 0
    all_warnings = []
    total_records = len(records)
    
    if args.verbose or total_records > 50:
        print(f"Validating {total_records} records...")
    
    for idx, record in enumerate(records, 1):
        # Show progress for large files
        if (args.verbose or total_records > 50) and (idx % 50 == 0 or idx == total_records):
            print(f"Validating {idx}/{total_records}...", end='\r')
        
        is_valid, warnings = validate_record(
            record, logger, 
            row_num=idx, 
            verbose=args.verbose,
            exclude_out_of_range=args.exclude_out_of_range
        )
        if is_valid:
            valid_records.append(record)
            all_warnings.extend(warnings)  # Range warnings
        else:
            skipped_count += 1
            serial_no = record.get('SerialNo', 'N/A')
            skip_reason = warnings[0] if warnings else "Validation failed"
            
            # Track skipped record for report
            skipped_records.append({
                'SerialNo': serial_no,
                'reason': skip_reason,
                'row': idx,
                'record': record
            })
            
            if args.verbose:
                logger.warning(f"Skipped invalid record (row {idx}): SerialNo={serial_no}, reason: {skip_reason}")
            else:
                logger.warning(f"Skipped invalid record: SerialNo={serial_no}, reason: {skip_reason}")
    
    if args.verbose or total_records > 50:
        print()  # New line after progress
    
    if not valid_records:
        logger.error("No valid records found after validation")
        print("Error: No valid records found after validation")
        return 1
    
    # Deduplicate records
    file_mtime = input_file.stat().st_mtime
    deduplicated = deduplicate_records(valid_records, file_mtime, logger)
    
    logger.info(f"Processing {len(deduplicated)} unique SerialNos")
    
    # Find pallet workbook
    workbook_path = find_pallet_workbook(args.xlsx, logger)
    if not workbook_path:
        logger.error("Could not find pallet workbook")
        print("Error: Could not find pallet workbook")
        return 1
    
    if args.dry_run:
        print(f"[DRY-RUN] Would update pallet workbook: {workbook_path.name}")
    else:
        print(f"Updating pallet workbook: {workbook_path.name}")
    logger.info(f"Updating pallet workbook: {workbook_path}")
    
    # Validate workbook structure
    is_valid, structure_warnings = validate_workbook_structure(workbook_path, logger)
    if not is_valid:
        logger.error("Workbook structure validation failed")
        print("Error: Workbook structure validation failed. Check logs for details.")
        return 1
    
    if structure_warnings:
        print("Warning: Workbook structure issues detected:")
        for warning in structure_warnings:
            print(f"  - {warning}")
        if not args.dry_run:
            response = input("Continue anyway? (y/n): ").strip().lower()
            if response != 'y':
                return 1
    
    # Create backup (unless dry-run)
    if not args.dry_run:
        backup_path = create_backup(workbook_path, logger)
        if not backup_path:
            logger.warning("Backup creation failed, but continuing...")
    else:
        logger.info("[DRY-RUN] Skipping backup creation")
        backup_path = None
    
    # Update workbook with error recovery
    updated_count = 0
    added_count = 0
    change_details = []
    update_errors = []
    
    try:
        updated_count, added_count, change_details = update_excel_workbook(
            workbook_path, deduplicated, logger, 
            dry_run=args.dry_run, 
            show_progress=(args.verbose or len(deduplicated) > 50)
        )
    except PermissionError:
        logger.error("Cannot update workbook - file may be open in Excel. Please close the file and try again.")
        print("Error: Cannot update workbook - file may be open in Excel.")
        print("Please close the Excel file and try again.")
        return 1
    except Exception as e:
        # Better error recovery: log error but continue if possible
        error_msg = f"Error updating workbook: {e}"
        logger.error(error_msg)
        update_errors.append(error_msg)
        
        # If we got partial results, use them
        if change_details:
            logger.warning("Partial update completed before error. Some records may have been updated.")
            # Extract counts from change_details
            updated_count = sum(1 for c in change_details if c.get('Action') == 'Updated')
            added_count = sum(1 for c in change_details if c.get('Action') == 'Added')
        else:
            # No partial results, this is a critical error
            print(f"Error: {error_msg}")
            print("No records were updated. Check logs for details.")
            return 1
    
    # Generate summary report (unless dry-run)
    report_path = None
    if not args.dry_run:
        report_path = generate_summary_report(
            change_details, skipped_records, logger,
            input_file.name, workbook_path.name
        )
        if report_path:
            logger.info(f"Summary report saved to: {report_path}")
    
    # Archive input file (only if update was successful and not dry-run)
    if not args.dry_run and updated_count + added_count > 0:
        archive_file(input_file, logger)
    elif args.dry_run:
        logger.info("[DRY-RUN] Skipping file archiving")
    elif updated_count + added_count == 0:
        logger.warning("No records were updated/added. File not archived.")
    
    # Log summary
    if args.verbose:
        all_serials = list(deduplicated.keys())
        logger.info(f"Import complete - Updated: {updated_count}, Added: {added_count}, Skipped: {skipped_count}")
        logger.info(f"All SerialNos processed ({len(all_serials)}): {', '.join(all_serials)}")
    else:
        first_10_serials = list(deduplicated.keys())[:10]
        logger.info(f"Import complete - Updated: {updated_count}, Added: {added_count}, Skipped: {skipped_count}")
        logger.info(f"First 10 SerialNos: {', '.join(first_10_serials)}")
    
    if all_warnings:
        logger.info(f"Total warnings: {len(all_warnings)}")
    
    logger.info("=" * 60)
    
    # Console output
    if args.dry_run:
        print("\n[DRY-RUN] Import preview complete.")
        print(f"Would update: {updated_count}")
        print(f"Would add: {added_count}")
        print(f"Skipped: {skipped_count}")
        print("\nNo changes were made to the workbook.")
    else:
        print("\nImport complete.")
        print(f"Updated: {updated_count}")
        print(f"Added: {added_count}")
        print(f"Skipped: {skipped_count}")
        if all_warnings:
            print(f"Warnings: {len(all_warnings)} (check log for details)")
        if update_errors:
            print(f"Errors: {len(update_errors)} (check log for details)")
        if report_path:
            print(f"\nSummary report: {report_path.name}")
        print("\nOpen the pallet sheet, scan barcodes, and print.")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())

