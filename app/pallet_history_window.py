#!/usr/bin/env python3
"""
Pallet History Window - View and manage completed pallets

Provides a window for viewing pallet history and opening exported files.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
from typing import Optional, List
from datetime import datetime, timedelta
import subprocess
import platform
import os
import sys

from app.pallet_manager import PalletManager
from app.path_utils import get_base_dir


class PalletHistoryWindow:
    """Window for viewing and managing pallet history"""
    
    def __init__(self, parent: tk.Tk, pallet_manager: PalletManager, customer_manager=None):
        """
        Initialize history window.
        
        Args:
            parent: Parent window (main GUI)
            pallet_manager: PalletManager instance
            customer_manager: Optional CustomerManager instance for customer filtering
        """
        self.parent = parent
        self.pallet_manager = pallet_manager
        self.customer_manager = customer_manager
        
        self.window = tk.Toplevel(parent)
        from app.version import get_version
        version = get_version()
        self.window.title(f"Pallet History - {version}")
        
        # Set window icon
        try:
            parent_icon = parent.iconbitmap()
            if parent_icon:
                self.window.iconbitmap(parent_icon)
        except:
            pass
        
        # Start with a reasonable size
        self.window.geometry("900x700")
        
        # Maximize the window (platform-specific)
        import platform
        try:
            if platform.system() == 'Windows':
                self.window.state('zoomed')
            elif platform.system() == 'Darwin':  # macOS
                # Manual maximization for macOS (more reliable than zoomed)
                self.window.update_idletasks()
                screen_width = self.window.winfo_screenwidth()
                screen_height = self.window.winfo_screenheight()
                # Leave small margin for dock/menu bar
                self.window.geometry(f"{screen_width-20}x{screen_height-100}+10+50")
            else:  # Linux
                try:
                    self.window.attributes('-zoomed', 1)
                except:
                    pass
        except Exception as e:
            print(f"Could not maximize History window: {e}")
            # Fallback to large fixed size if maximization fails
            self.window.geometry("1200x900+50+50")
        
        self.selected_pallet: Optional[dict] = None
        self.checkbox_states: dict = {}  # Track checkbox states by tree item id
        self.item_to_pallet: dict = {}  # Map tree item id -> pallet record
        
        self.setup_ui()
        self.load_history()

        # Bring window to front and focus it
        self.window.lift()
        self.window.focus_force()
        self.window.attributes('-topmost', True)  # Keep on top briefly
        self.window.after(100, lambda: self.window.attributes('-topmost', False))  # Remove topmost after a moment
    
    def setup_ui(self):
        """Create UI layout"""
        # Header
        header = tk.Frame(self.window)
        header.pack(fill=tk.X, padx=10, pady=5)
        
        tk.Label(header, text="📋 Pallet History", 
                font=("Arial", 14, "bold")).pack(side=tk.LEFT)
        
        tk.Button(header, text="Close", command=self.window.destroy).pack(side=tk.RIGHT)
        
        # Filter container (holds filter row and select all checkbox)
        filter_container = tk.Frame(self.window)
        filter_container.pack(fill=tk.X, padx=10, pady=5)
        
        # Filter frame (all filters in one row)
        filter_frame = tk.Frame(filter_container)
        filter_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Date filter
        date_filter_frame = tk.Frame(filter_frame)
        date_filter_frame.pack(side=tk.LEFT, padx=5)
        
        tk.Label(date_filter_frame, text="Date Filter:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        self.filter_var = tk.StringVar(value="Today")
        filter_options = [
            ("All", "All"),
            ("Today", "Today"),
            ("This Week", "This Week"),
            ("This Month", "This Month"),
            ("This Year", "This Year")
        ]
        
        for text, value in filter_options:
            tk.Radiobutton(
                date_filter_frame, 
                text=text, 
                variable=self.filter_var, 
                value=value, 
                command=self.load_history,
                font=("Arial", 9)
            ).pack(side=tk.LEFT, padx=3)
        
        # Customer filter
        customer_filter_frame = tk.Frame(filter_frame)
        customer_filter_frame.pack(side=tk.LEFT, padx=10)
        
        tk.Label(customer_filter_frame, text="Customer:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        self.customer_filter_var = tk.StringVar(value="ALL")
        # Don't use command parameter - it causes infinite loop. Use trace instead.
        self.customer_filter_menu = tk.OptionMenu(customer_filter_frame, self.customer_filter_var, "ALL")
        self.customer_filter_menu.config(font=("Arial", 9), width=25)
        self.customer_filter_menu.pack(side=tk.LEFT, padx=3)
        
        # Add trace to detect changes (avoids infinite loop from command parameter)
        self.customer_filter_var.trace('w', self._on_customer_filter_changed)
        
        # Update customer filter options
        self._update_customer_filter_options()
        
        # Barcode search
        search_frame = tk.Frame(filter_frame)
        search_frame.pack(side=tk.LEFT, padx=10)
        
        tk.Label(search_frame, text="Search Barcode:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self._on_search_changed)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, font=("Arial", 9), width=20)
        search_entry.pack(side=tk.LEFT, padx=3)
        
        # Select All checkbox (below filters, aligned left)
        select_all_frame = tk.Frame(filter_container)
        select_all_frame.pack(fill=tk.X, padx=5)
        
        self.select_all_var = tk.BooleanVar(value=False)
        select_all_checkbox = tk.Checkbutton(
            select_all_frame, 
            text="Select All", 
            variable=self.select_all_var,
            command=self.on_select_all_clicked,
            font=("Arial", 10, "bold")
        )
        select_all_checkbox.pack(side=tk.LEFT, padx=0)
        
        # Main content area (split into table and details)
        main_frame = tk.Frame(self.window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        # Left side: Table
        table_frame = tk.LabelFrame(main_frame, text="Pallets", font=("Arial", 10, "bold"))
        table_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Treeview for table with Windows File Explorer-style selection
        columns = ("Select", "Pallet #", "Completed", "File Name")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=15, selectmode="extended")
        
        # Configure columns
        # Header checkbox - clicking it selects/deselects all
        self.header_select_all = False  # Track header checkbox state
        self.tree.heading("Select", text="☐", command=self.on_header_checkbox_click)
        self.tree.column("Select", width=50, anchor="center")
        
        # Sortable headers
        self.sort_column = None  # Track current sort column
        self.sort_reverse = False  # Track sort direction
        self.tree.heading("Pallet #", text="Pallet #", command=lambda: self._sort_by_column("Pallet #"))
        self.tree.column("Pallet #", width=100)
        self.tree.heading("Completed", text="Completed", command=lambda: self._sort_by_column("Completed"))
        self.tree.column("Completed", width=150)
        self.tree.heading("File Name", text="File Name", command=lambda: self._sort_by_column("File Name"))
        self.tree.column("File Name", width=200)
        
        # Bind click events for checkbox-based selection
        self.tree.bind("<Button-1>", self.on_tree_click)
        
        # Configure selection tags for visual highlighting
        self.tree.tag_configure("selected", background="#0078d4", foreground="white")
        self.tree.tag_configure("unselected", background="white", foreground="black")
        
        self.tree.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar for table
        table_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", 
                                       command=self.tree.yview)
        self.tree.configure(yscrollcommand=table_scrollbar.set)
        table_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Right side: Details
        details_frame = tk.LabelFrame(main_frame, text="Pallet Details", 
                                      font=("Arial", 10, "bold"))
        details_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        # Details text area
        self.details_text = tk.Text(details_frame, wrap=tk.WORD, height=15, 
                                    font=("Courier", 9))
        self.details_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Action buttons
        action_frame = tk.Frame(details_frame)
        action_frame.pack(fill=tk.X, padx=5, pady=5)
        
        tk.Button(action_frame, text="Open Export File", 
                 command=self.open_export_file, width=15).pack(side=tk.LEFT, padx=2)
        tk.Button(action_frame, text="Open Export Folder", 
                 command=self.open_export_folder, width=15).pack(side=tk.LEFT, padx=2)
        tk.Button(action_frame, text="Reset Pallet",
                 command=self.reset_selected_pallet, width=15,
                 bg="#ff9800", fg="black", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=2)
        tk.Button(action_frame, text="Delete Pallet",
                 command=self.delete_selected_pallet, width=15,
                 bg="#f44336", fg="white", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=2)
        
        # Multi-select actions
        multi_action_frame = tk.Frame(details_frame)
        multi_action_frame.pack(fill=tk.X, padx=5, pady=5)
        
        tk.Label(multi_action_frame, text="Multi-Select:", 
                font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=5)
        tk.Button(multi_action_frame, text="Create PDF & Print", 
                 command=self.create_pdf_and_print, width=18, 
                 bg="#2196F3", fg="black", font=("Arial", 9)).pack(side=tk.LEFT, padx=2)
    
    def load_history(self):
        """Load and display pallet history"""
        # Clear existing items (fast operation)
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Show loading indicator
        loading_item = self.tree.insert("", tk.END, values=("Loading...", "", ""))
        
        # Defer heavy operations to keep UI responsive
        def load_and_populate():
            try:
                # Ensure data is loaded (might be deferred)
                if not self.pallet_manager.data.get('pallets') and hasattr(self.pallet_manager, 'load_history'):
                    self.pallet_manager.data = self.pallet_manager.load_history()
                
                # Get filters
                filter_value = self.filter_var.get()
                customer_filter = self.customer_filter_var.get()
                search_term = self.search_var.get().strip().upper() if hasattr(self, 'search_var') else ""
                
                # Get all pallets (use cached data directly)
                all_pallets = self.pallet_manager.data.get('pallets', [])
                
                # Filter out pallets whose exported files no longer exist
                project_root = get_base_dir()
                pallets_dir = project_root / "PALLETS"
                valid_pallets = []
                for pallet in all_pallets:
                    exported_file = pallet.get('exported_file', '')
                    if exported_file:
                        # Check if the exported file exists
                        file_path = Path(exported_file)
                        # Handle both absolute and relative paths
                        if file_path.is_absolute():
                            # Absolute path - check directly
                            if not file_path.exists():
                                continue  # Skip this pallet
                        else:
                            # Relative path - try multiple locations
                            # First try: relative to PALLETS directory
                            full_path = pallets_dir / file_path
                            if not full_path.exists():
                                # Try: just the filename in PALLETS directory
                                filename_only = Path(file_path).name
                                full_path = pallets_dir / filename_only
                                if not full_path.exists():
                                    # Try: search in date subdirectories
                                    found = False
                                    if pallets_dir.exists():
                                        try:
                                            for date_dir in pallets_dir.iterdir():
                                                if date_dir.is_dir():
                                                    potential_path = date_dir / filename_only
                                                    if potential_path.exists():
                                                        found = True
                                                        break
                                        except (PermissionError, OSError) as e:
                                            # Can't read directory, skip search
                                            print(f"Warning: Could not search PALLETS directory: {e}")
                                    if not found:
                                        continue  # Skip this pallet - file doesn't exist
                        valid_pallets.append(pallet)
                    else:
                        # Keep pallets that haven't been exported yet
                        valid_pallets.append(pallet)
                
                # Filter pallets based on selected time period
                pallets = self._filter_pallets_by_date(valid_pallets, filter_value)
                
                # Filter by customer
                if customer_filter != "ALL":
                    filtered_pallets = []
                    for p in pallets:
                        customer_info = p.get('customer', {})
                        if customer_info:
                            # Check both display_name and fallback to name|business format
                            display_name = customer_info.get('display_name')
                            if not display_name:
                                name = customer_info.get('name', '')
                                business = customer_info.get('business', '')
                                display_name = f"{name} | {business}" if name and business else None
                            
                            if display_name == customer_filter:
                                filtered_pallets.append(p)
                    pallets = filtered_pallets
                
                # Filter by barcode search
                if search_term:
                    matching_pallets = []
                    for pallet in pallets:
                        serials = pallet.get('serial_numbers', [])
                        # Exact serial match only (case-insensitive) to avoid false positives.
                        if any(search_term == str(serial).strip().upper() for serial in serials):
                            matching_pallets.append(pallet)
                    pallets = matching_pallets
                
                # Sort by pallet_number descending (most recent first)
                pallets.sort(key=lambda x: x.get('pallet_number', 0), reverse=True)
                
                # Remove loading indicator
                self.tree.delete(loading_item)
                
                # Show search result message if searching
                if search_term:
                    if not pallets:
                        # Show message in tree temporarily
                        self.tree.insert("", tk.END, values=("", f"No pallets found with barcode '{search_term}'", "", ""))
                    else:
                        # Show count in status (if we had a status bar, but for now just populate)
                        pass
                
                # Populate table in chunks (non-blocking)
                self._populate_table(pallets)
                
                # Auto-select if only one pallet remains from search
                if search_term and len(pallets) == 1:
                    # Wait a moment for table to populate, then auto-select
                    # Capture pallet in closure to avoid lambda capture issues
                    pallet_to_select = pallets[0]
                    self.window.after(100, lambda p=pallet_to_select: self._auto_select_single_pallet(p))
            except Exception as e:
                # Remove loading indicator on error
                try:
                    self.tree.delete(loading_item)
                except:
                    pass
                # Show error
                messagebox.showerror("Error", f"Failed to load history: {e}", parent=self.window)
        
        # Schedule loading after UI is ready
        self.window.after(10, load_and_populate)
    
    def _filter_pallets_by_date(self, pallets: List[dict], filter_value: str) -> List[dict]:
        """
        Filter pallets by date based on filter selection.
        
        Args:
            pallets: List of all pallet dictionaries
            filter_value: Filter selection ("All", "Today", "This Week", "This Month", "This Year")
        
        Returns:
            Filtered list of pallets
        """
        if filter_value == "All":
            return pallets
        
        now = datetime.now()
        filtered = []
        
        for pallet in pallets:
            completed_at = pallet.get('completed_at', '')
            if not completed_at:
                continue
            
            try:
                # Parse completed_at timestamp (format: "YYYY-MM-DD HH:MM:SS")
                pallet_date = datetime.strptime(completed_at.split()[0], "%Y-%m-%d")
            except (ValueError, IndexError):
                # Skip if date parsing fails
                continue
            
            # Calculate days difference
            days_diff = (now.date() - pallet_date.date()).days
            
            if filter_value == "Today":
                # Show only pallets from today
                if days_diff == 0:
                    filtered.append(pallet)
            
            elif filter_value == "This Week":
                # Show pallets from the last 7 days (including today)
                if 0 <= days_diff <= 6:
                    filtered.append(pallet)
            
            elif filter_value == "This Month":
                # Show pallets from current month
                if pallet_date.year == now.year and pallet_date.month == now.month:
                    filtered.append(pallet)
            
            elif filter_value == "This Year":
                # Show pallets from current year
                if pallet_date.year == now.year:
                    filtered.append(pallet)
        
        return filtered
    
    def _populate_table(self, pallets):
        """Populate table with pallets (chunked for responsiveness)"""
        if not pallets:
            return
        
        # Clear checkbox states for new data
        self.checkbox_states = {}
        self.item_to_pallet = {}
        self.header_select_all = False  # Reset header checkbox
        if hasattr(self, 'select_all_var'):
            self.select_all_var.set(False)  # Reset select all
        
        # Update header text to unchecked
        self.tree.heading("Select", text="☐")
        
        # Process in chunks to keep UI responsive
        chunk_size = 20  # Insert 20 rows at a time
        current_index = [0]  # Use list to allow modification in nested function
        
        def insert_next_chunk():
            """Insert next chunk of rows"""
            end_index = min(current_index[0] + chunk_size, len(pallets))
            
            for i in range(current_index[0], end_index):
                pallet = pallets[i]
                pallet_num = pallet.get('pallet_number', 'N/A')
                completed = pallet.get('completed_at', 'N/A')
                exported_file = pallet.get('exported_file', '')
                file_name = Path(exported_file).name if exported_file else 'N/A'

                # Add reset indicator to file name
                if pallet.get('reset', False):
                    file_name = f"[RESET] {file_name}"
                
                # Insert with checkbox column (empty checkbox symbol) and tags for selection highlighting
                item_id = self.tree.insert(
                    "", tk.END,
                    values=("☐", pallet_num, completed, file_name),
                    tags=(str(pallet_num), "unselected")
                )
                self.checkbox_states[item_id] = False
                self.item_to_pallet[item_id] = pallet
            
            current_index[0] = end_index
            
            # Schedule next chunk if more rows to process
            if current_index[0] < len(pallets):
                self.window.after(10, insert_next_chunk)
        
        # Start inserting (non-blocking)
        self.window.after(10, insert_next_chunk)
    
    def on_tree_click(self, event):
        """Handle clicks on tree - checkbox-based multi-selection"""
        item = self.tree.identify_row(event.y)
        if not item:
            return

        values = list(self.tree.item(item, 'values'))
        if len(values) <= 1:
            return

        # Get column clicked
        column = self.tree.identify_column(event.x)
        column_index = int(column.replace("#", "")) - 1  # Convert "#1" to 0, "#2" to 1, etc.

        # Only handle checkbox column clicks (column 0)
        if column_index == 0:
            # Toggle checkbox state (independent of other checkboxes)
            current_state = self.checkbox_states.get(item, False)
            new_state = not current_state
            self.checkbox_states[item] = new_state

            # Update checkbox display
            self._update_checkbox_display(item)

            # Sync Treeview selection with checkbox state
            if new_state:
                self.tree.selection_add(item)
            else:
                self.tree.selection_remove(item)

            # Update select all checkbox state
            self._update_select_all_state()

            # Show details based on selection
            self._update_details_for_selection()
    
    
    def _select_single_item(self, item, pallet_num):
        """Select a single item and deselect all others (Windows File Explorer style)"""
        # Clear all selections first
        self.tree.selection_set(())  # Clear all
        
        # Select only this item
        self.tree.selection_add(item)
        
        # Update all checkbox states - only this one is checked
        for child in self.tree.get_children():
            self.checkbox_states[child] = (child == item)
            self._update_checkbox_display(child)
    
    def _sync_checkbox_with_selection(self):
        """Sync checkbox states with Treeview selection"""
        selected_items = self.tree.selection()
        
        # Update checkbox states based on selection
        for child in self.tree.get_children():
            is_selected = child in selected_items
            self.checkbox_states[child] = is_selected
            self._update_checkbox_display(child)
        
        # Update select all checkbox state
        self._update_select_all_state()
    
    def _update_checkbox_display(self, item):
        """Update checkbox display for a specific item and sync visual highlighting"""
        try:
            values = list(self.tree.item(item, 'values'))
            if len(values) > 0:
                # Update checkbox symbol
                is_checked = self.checkbox_states.get(item, False)
                values[0] = "☑" if is_checked else "☐"
                self.tree.item(item, values=values)
                
                # Update visual highlighting tags (Windows File Explorer style)
                current_tags = list(self.tree.item(item, 'tags'))
                # Remove old selection tags
                current_tags = [tag for tag in current_tags if tag not in ['selected', 'unselected']]
                # Add appropriate tag
                if is_checked:
                    current_tags.append('selected')
                else:
                    current_tags.append('unselected')
                self.tree.item(item, tags=current_tags)
        except Exception as e:
            print(f"DEBUG: Error updating checkbox: {e}")  # Debug output
    
    def _show_pallet_details(self, item_id):
        """Show details for a pallet (deferred to avoid lag)"""
        self.selected_pallet = self.item_to_pallet.get(item_id)
        
        if self.selected_pallet:
            self.update_details()
    
    def _update_details_for_selection(self):
        """Update details panel based on current checkbox selection"""
        selected_count = sum(1 for checked in self.checkbox_states.values() if checked)
        
        if selected_count == 0:
            # No selection - clear details
            self.selected_pallet = None
            self.details_text.delete(1.0, tk.END)
            self.details_text.insert(1.0, "No pallet selected.\n\nClick on a pallet row to view details.")
        elif selected_count == 1:
            # Exactly one selected - show its details
            selected_items = [
                item_id for item_id, checked in self.checkbox_states.items()
                if checked
            ]
            if selected_items:
                self._show_pallet_details(selected_items[0])
        else:
            # Multiple selected - show summary
            self.selected_pallet = None
            self.details_text.delete(1.0, tk.END)
            self.details_text.insert(1.0, f"{selected_count} pallets selected.\n\nUse multi-select actions below.")
    
    def on_header_checkbox_click(self):
        """Handle header checkbox click - toggles select all/deselect all"""
        # Toggle state
        self.header_select_all = not self.header_select_all
        
        # Update header text
        header_text = "☑" if self.header_select_all else "☐"
        self.tree.heading("Select", text=header_text)
        
        # Update all checkbox states
        self._update_all_checkboxes(self.header_select_all)
        
        # Update the separate "Select All" checkbox to match
        if hasattr(self, 'select_all_var'):
            self.select_all_var.set(self.header_select_all)
    
    def on_select_all_clicked(self):
        """Handle Select All checkbox click (separate checkbox)"""
        select_all = self.select_all_var.get()
        print(f"DEBUG: Select All clicked - State: {select_all}")  # Debug output
        
        # Update header checkbox state
        self.header_select_all = select_all
        header_text = "☑" if self.header_select_all else "☐"
        self.tree.heading("Select", text=header_text)
        
        # Update all checkbox states immediately
        self._update_all_checkboxes(select_all)
    
    def _update_all_checkboxes(self, checked):
        """Update all checkboxes in the table and sync Treeview selection"""
        if checked:
            # Select all items in Treeview
            all_items = list(self.tree.get_children())
            self.tree.selection_set(all_items)
        else:
            # Deselect all items in Treeview
            self.tree.selection_set(())
        
        # Update checkbox states and display
        for item in self.tree.get_children():
            self.checkbox_states[item] = checked
            self._update_checkbox_display(item)
        
        # Update details panel based on new selection
        self._update_details_for_selection()
    
    def _update_select_all_state(self):
        """Update Select All checkbox and header checkbox based on current selection"""
        # Check if all items are selected
        all_items = list(self.tree.get_children())
        if not all_items:
            self.header_select_all = False
            if hasattr(self, 'select_all_var'):
                self.select_all_var.set(False)
            self.tree.heading("Select", text="☐")
            return
        
        all_checked = all(
            self.checkbox_states.get(item, False)
            for item in all_items
        )
        
        # Update header checkbox state and text
        self.header_select_all = all_checked
        header_text = "☑" if all_checked else "☐"
        self.tree.heading("Select", text=header_text)
        
        # Update select all checkbox (without triggering event)
        if hasattr(self, 'select_all_var'):
            self.select_all_var.set(all_checked)
    
    def update_details(self):
        """Update details panel with selected pallet information"""
        if not self.selected_pallet:
            return
        
        self.details_text.delete(1.0, tk.END)
        
        pallet = self.selected_pallet
        details = f"Pallet #{pallet.get('pallet_number', 'N/A')}\n"
        details += f"Completed: {pallet.get('completed_at', 'N/A')}\n"
        
        # Show customer information if available
        customer_info = pallet.get('customer')
        if customer_info:
            customer_display = customer_info.get('display_name', f"{customer_info.get('name', 'N/A')} | {customer_info.get('business', 'N/A')}")
            details += f"Customer: {customer_display}\n"
        else:
            details += f"Customer: Not specified\n"
        
        # Show reset status
        if pallet.get('reset', False):
            details += f"Status: RESET - {pallet.get('reset_reason', 'Manual reset')}\n"
            details += f"Reset Date: {pallet.get('reset_at', 'N/A')}\n"
            details += f"Original Status: {'Exported' if pallet.get('exported_file') else 'Not Exported'}\n"
        else:
            details += f"Status: {'Exported' if pallet.get('exported_file') else 'Not Exported'}\n"

        details += f"Export File: {Path(pallet.get('exported_file', '')).name if pallet.get('exported_file') else 'N/A'}\n\n"
        details += "Serial Numbers:\n"
        details += "-" * 40 + "\n"
        
        serials = pallet.get('serial_numbers', [])
        for i, serial in enumerate(serials, start=1):
            details += f"{i:2d}. {serial}\n"
        
        # Show empty slots - support both 25 and 26 panel pallets
        # Excel template supports up to 26 panels (rows 5-30)
        max_slots = 26  # Maximum supported by Excel template
        if len(serials) < max_slots:
            for i in range(len(serials) + 1, max_slots + 1):
                details += f"{i:2d}. (empty)\n"
        
        self.details_text.insert(1.0, details)

    def reset_selected_pallet(self):
        """Reset the selected pallet, allowing its panels to be rescanned"""
        # Get selected pallets from checkboxes
        selected_pallets = self.get_selected_pallets()

        if not selected_pallets:
            # Fall back to single selected pallet if no checkboxes selected
            if not self.selected_pallet:
                messagebox.showwarning("No Selection", "Please select a pallet to reset.",
                                     parent=self.window)
                return
            selected_pallets = [self.selected_pallet]

        # For now, only allow resetting one pallet at a time
        if len(selected_pallets) > 1:
            messagebox.showwarning("Multiple Selection",
                                 "Please select only one pallet to reset at a time.",
                                 parent=self.window)
            return

        pallet = selected_pallets[0]
        pallet_num = pallet.get('pallet_number', 'N/A')
        exported_file = pallet.get('exported_file', '')

        # Check if pallet is already reset
        if pallet.get('reset', False):
            messagebox.showinfo("Already Reset",
                               f"Pallet #{pallet_num} has already been reset.",
                               parent=self.window)
            return

        # Confirm reset
        confirm_msg = f"Are you sure you want to reset Pallet #{pallet_num}?\n\n"
        confirm_msg += "This will:\n"
        confirm_msg += "• Keep the pallet in history with a reset record\n"
        confirm_msg += "• Allow the panels to be scanned again onto another pallet\n"
        confirm_msg += "• Preserve the export file (if it exists)\n\n"
        confirm_msg += "The pallet will remain visible in history but marked as reset.\n"
        confirm_msg += "Use 'Delete Pallet' if you want to remove it completely."

        if not messagebox.askyesno("Confirm Reset", confirm_msg, parent=self.window):
            return

        # Ask for reset reason
        from tkinter import simpledialog
        reason = simpledialog.askstring("Reset Reason",
                                       "Enter reason for resetting this pallet:",
                                       parent=self.window)
        if reason is None:  # User cancelled
            return
        if not reason.strip():
            reason = "Manual reset"

        # Reset the pallet
        try:
            success = self.pallet_manager.reset_pallet(pallet_num, reason)
            if success:
                messagebox.showinfo("Reset Successful",
                                  f"Pallet #{pallet_num} has been reset.\n\n"
                                  f"The panels from this pallet can now be scanned again.\n\n"
                                  f"Reason: {reason}",
                                  parent=self.window)

                # Clear selection and refresh history
                self.selected_pallet = None
                self.checkbox_states = {}
                self.load_history()
            else:
                messagebox.showerror("Reset Failed",
                                   f"Could not find Pallet #{pallet_num} in history.",
                                   parent=self.window)
        except Exception as e:
            messagebox.showerror("Reset Error",
                               f"An error occurred while resetting the pallet:\n{e}",
                               parent=self.window)

    def delete_selected_pallet(self):
        """Delete the selected pallet and its export file"""
        # Get selected pallets from checkboxes
        selected_pallets = self.get_selected_pallets()
        
        if not selected_pallets:
            # Fall back to single selected pallet if no checkboxes selected
            if not self.selected_pallet:
                messagebox.showwarning("No Selection", "Please select a pallet to delete.", 
                                     parent=self.window)
                return
            selected_pallets = [self.selected_pallet]
        
        # For now, only allow deleting one pallet at a time
        if len(selected_pallets) > 1:
            messagebox.showwarning("Multiple Selection", 
                                 "Please select only one pallet to delete at a time.", 
                                 parent=self.window)
            return
        
        pallet = selected_pallets[0]
        pallet_num = pallet.get('pallet_number', 'N/A')
        exported_file = pallet.get('exported_file', '')
        
        # Confirm deletion
        confirm_msg = f"Are you sure you want to delete Pallet #{pallet_num}?\n\n"
        confirm_msg += "This will:\n"
        confirm_msg += "• Delete the export file (if it exists)\n"
        confirm_msg += "• Remove the pallet from history\n"
        confirm_msg += "• Allow the panels to be scanned again onto another pallet\n\n"
        confirm_msg += "This action cannot be undone."
        
        if not messagebox.askyesno("Confirm Delete", confirm_msg, parent=self.window):
            return
        
        # Delete the physical file if it exists
        if exported_file:
            try:
                file_path = Path(exported_file)
                if not file_path.is_absolute():
                    # Make relative to project root
                    project_root = get_base_dir()
                    file_path = project_root / file_path
                
                # Also try searching in date subdirectories if direct path doesn't exist
                if not file_path.exists():
                    pallets_dir = project_root / "PALLETS"
                    filename = Path(exported_file).name
                    if pallets_dir.exists():
                        try:
                            for date_dir in pallets_dir.iterdir():
                                if date_dir.is_dir():
                                    potential_path = date_dir / filename
                                    if potential_path.exists():
                                        file_path = potential_path
                                        break
                        except (PermissionError, OSError) as e:
                            print(f"Warning: Could not search PALLETS directory for delete: {e}")
                
                if file_path.exists():
                    file_path.unlink()
                    print(f"Deleted export file: {file_path}")
            except Exception as e:
                # File deletion failed, but continue with pallet deletion
                print(f"Warning: Could not delete export file: {e}")
                messagebox.showwarning("File Deletion Warning", 
                                     f"Could not delete export file:\n{file_path}\n\n"
                                     f"Error: {e}\n\n"
                                     f"The pallet will still be removed from history.",
                                     parent=self.window)
        
        # Remove pallet from history
        try:
            success = self.pallet_manager.delete_pallet(pallet_num)
            if success:
                messagebox.showinfo("Delete Successful", 
                                  f"Pallet #{pallet_num} has been deleted.\n\n"
                                  f"The panels from this pallet can now be scanned again.",
                                  parent=self.window)
                
                # Clear selection and refresh history
                self.selected_pallet = None
                self.checkbox_states = {}
                self.load_history()
            else:
                messagebox.showerror("Delete Failed", 
                                   f"Could not find Pallet #{pallet_num} in history.",
                                   parent=self.window)
        except Exception as e:
            messagebox.showerror("Delete Error", 
                               f"An error occurred while deleting the pallet:\n{e}",
                               parent=self.window)
    
    def open_export_file(self):
        """Open the selected pallet's export file"""
        print("DEBUG: Open Export File button clicked")  # Debug output
        
        # Get selected pallets from checkboxes
        selected_pallets = self.get_selected_pallets()
        
        if not selected_pallets:
            # Fall back to single selected pallet if no checkboxes selected
            if not self.selected_pallet:
                messagebox.showwarning("No Selection", "Please select a pallet first.", 
                                     parent=self.window)
                return
            selected_pallets = [self.selected_pallet]
        
        # Use first selected pallet
        pallet = selected_pallets[0]
        exported_file = pallet.get('exported_file')
        if not exported_file:
            messagebox.showwarning("No Export File", 
                                 "This pallet has no export file.", 
                                 parent=self.window)
            return
        
        file_path = Path(exported_file)
        if not file_path.is_absolute():
            # Make relative to project root using proper path utilities
            project_root = get_base_dir()
            file_path = project_root / file_path
        
        if not file_path.exists():
            messagebox.showerror(
                "File Not Found",
                f"Export file not found:\n{file_path}\n\n"
                "The file may have been moved or deleted.",
                parent=self.window
            )
            return
        
        try:
            system = platform.system()
            if system == 'Windows':
                subprocess.run(['explorer', str(file_path.absolute())], check=False)
            elif system == 'Darwin':  # macOS
                subprocess.run(['open', str(file_path.absolute())], check=False)
            else:  # Linux
                subprocess.run(['xdg-open', str(file_path.absolute())], check=False)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file:\n{e}", 
                               parent=self.window)
    
    def open_export_folder(self):
        print("DEBUG: Open Export Folder button clicked")  # Debug output
        """Open the export folder"""
        # Use proper path utilities for packaged/development environments
        project_root = get_base_dir()
        export_dir = project_root / "PALLETS"
        
        # Create folder if it doesn't exist (it should, but be helpful)
        if not export_dir.exists():
            try:
                export_dir.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                messagebox.showerror("Error", 
                                   f"Export folder does not exist and could not be created:\n{e}\n\n"
                                   f"Expected location: {export_dir}",
                                   parent=self.window)
                return
        
        try:
            system = platform.system()
            if system == 'Windows':
                subprocess.run(['explorer', str(export_dir.absolute())], check=False)
            elif system == 'Darwin':  # macOS
                subprocess.run(['open', str(export_dir.absolute())], check=False)
            else:  # Linux
                subprocess.run(['xdg-open', str(export_dir.absolute())], check=False)
        except Exception as e:
            messagebox.showerror("Error", f"Could not open folder:\n{e}", 
                               parent=self.window)
    
    def get_selected_pallets(self) -> List[dict]:
        """Get list of selected pallet dicts from checkbox states"""
        # Get selected pallet numbers from checkbox states
        selected_pallet_nums = [
            item_id for item_id, checked in self.checkbox_states.items()
            if checked
        ]
        
        if not selected_pallet_nums:
            return []
        
        selected_pallets = []
        for item_id in selected_pallet_nums:
            pallet = self.item_to_pallet.get(item_id)
            if pallet:
                selected_pallets.append(pallet)
        
        print(f"DEBUG: Selected pallets: {len(selected_pallets)}")  # Debug output
        return selected_pallets
    
    def create_pdf_and_print(self):
        """Create PDF from selected Excel files and print
        
        Optimized for low-end systems with comprehensive error handling
        """
        from app.debug_logger import get_logger
        
        logger = get_logger()
        logger.section("PDF Export & Print Started")
        logger.start_timer("pdf_export_total")
        logger.log_memory_usage()
        
        print("DEBUG: Create PDF & Print button clicked")  # Debug output
        selected_pallets = self.get_selected_pallets()
        
        logger.info(f"Selected {len(selected_pallets)} pallets")
        
        if not selected_pallets:
            logger.warning("No pallets selected")
            messagebox.showwarning("No Selection", 
                                 "Please select one or more pallets first.\n\n"
                                 "Click the checkboxes to select pallets.",
                                 parent=self.window)
            return
        
        # Filter to only exported pallets
        exported_pallets = [p for p in selected_pallets if p.get('exported_file')]
        
        if not exported_pallets:
            messagebox.showwarning("No Export Files", 
                                 "Selected pallets have no export files.",
                                 parent=self.window)
            return
        
        # Get file paths using proper path utilities
        project_root = get_base_dir()
        file_paths = []
        for pallet in exported_pallets:
            exported_file = pallet.get('exported_file')
            if exported_file:
                file_path = Path(exported_file)
                if not file_path.is_absolute():
                    file_path = project_root / file_path
                if file_path.exists():
                    file_paths.append(file_path)
        
        if not file_paths:
            messagebox.showerror("Error", 
                               "No export files found for selected pallets.",
                               parent=self.window)
            return
        
        try:
            # Check for reportlab (required for PDF creation)
            try:
                from reportlab.lib.pagesizes import letter, landscape
                from reportlab.pdfgen import canvas
            except ImportError as e:
                # reportlab is required - show error and instructions
                import sys
                import os
                
                # Get more detailed error info for debugging
                error_details = str(e)
                is_packaged = getattr(sys, 'frozen', False)
                python_path = sys.executable
                
                # Check if reportlab module exists but import failed
                try:
                    import reportlab
                    reportlab_found = True
                    reportlab_path = os.path.dirname(reportlab.__file__) if hasattr(reportlab, '__file__') else "unknown"
                except ImportError:
                    reportlab_found = False
                    reportlab_path = "not found"
                
                error_msg = (
                    f"PDF creation requires 'reportlab' package.\n\n"
                    f"Error: {error_details}\n\n"
                    f"Debug Info:\n"
                    f"- Packaged app: {is_packaged}\n"
                    f"- Python path: {python_path}\n"
                    f"- Reportlab found: {reportlab_found}\n"
                    f"- Reportlab path: {reportlab_path}\n\n"
                    f"If this is a packaged .exe file:\n"
                    f"Reportlab was not included during build.\n"
                    f"You need to rebuild the executable after installing reportlab:\n\n"
                    f"1. Install reportlab: pip install reportlab\n"
                    f"2. Rebuild: scripts\\rebuild_exe.bat\n\n"
                    f"If running from source:\n"
                    f"Install reportlab: pip install reportlab"
                )
                
                messagebox.showerror(
                    "Missing Dependency - reportlab",
                    error_msg,
                    parent=self.window
                )
                return
            
            # Automatically determine PDF filename and location
            # Use the same directory as the first Excel file
            first_file_dir = file_paths[0].parent
            
            if len(file_paths) == 1:
                default_name = file_paths[0].stem + ".pdf"
            else:
                default_name = f"Pallets_{len(file_paths)}_combined.pdf"
            
            pdf_path = first_file_dir / default_name
            
            # If PDF already exists, append a number to make it unique
            if pdf_path.exists():
                base_name = pdf_path.stem
                counter = 1
                while pdf_path.exists():
                    pdf_path = first_file_dir / f"{base_name}_{counter}.pdf"
                    counter += 1
                    if counter > 1000:  # Safety limit
                        messagebox.showerror("Error", 
                                           "Too many PDF files with same name. Please clean up.",
                                           parent=self.window)
                        return
            
            # Show progress
            progress_window = tk.Toplevel(self.window)
            progress_window.title("Creating PDF...")
            progress_window.geometry("300x100")
            progress_window.transient(self.window)
            progress_window.grab_set()
            
            progress_label = tk.Label(progress_window, text="Creating PDF...", 
                                     font=("Arial", 10))
            progress_label.pack(pady=20)
            progress_window.update()
            
            # Create PDF from Excel files
            # Returns list of PDF paths (one per pallet)
            individual_pdfs = self._excel_to_pdf(file_paths, pdf_path, progress_label)
            
            # If multiple PDFs, create a merged version for printing
            if len(individual_pdfs) > 1:
                progress_label.config(text="Creating merged PDF for printing...")
                progress_label.master.update()

                # Create PRINT folder inside HISTORY
                print_folder = pdf_path.parent / "PRINT"
                print_folder.mkdir(exist_ok=True)

                # Create merged PDF filename with timestamp
                timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
                merged_pdf_name = f"PRINT_Combined_{timestamp}.pdf"
                merged_pdf_path = print_folder / merged_pdf_name

                # Merge all individual PDFs
                logger.start_timer("merge_pdfs")
                self._merge_pdfs([str(p) for p in individual_pdfs], merged_pdf_path)
                logger.end_timer("merge_pdfs")
                logger.info(f"Merged PDF created: {merged_pdf_path}")

                progress_window.destroy()

                # Show success message
                messagebox.showinfo(
                    "PDFs Created Successfully",
                    f"✓ Individual PDFs saved: {len(individual_pdfs)} files\n"
                    f"  Location: {pdf_path.parent}\n\n"
                    f"✓ Merged print PDF created:\n"
                    f"  {merged_pdf_path.name}\n\n"
                    f"Opening print dialog for merged PDF...",
                    parent=self.window
                )

                # Open print dialog for merged PDF only
                logger.info("Opening print dialog for merged PDF")
                self._print_pdf(merged_pdf_path)
            else:
                progress_window.destroy()

                # Single PDF - just print it
                logger.info(f"Opening print dialog for single PDF: {individual_pdfs[0]}")
                self._print_pdf(individual_pdfs[0])
            
            logger.end_timer("pdf_export_total")
            logger.log_memory_usage()
            logger.info("PDF export and print completed successfully")
            
        except Exception as e:
            logger.error(f"Failed to create PDF: {e}", exc_info=e)
            logger.end_timer("pdf_export_total")
            
            if 'progress_window' in locals():
                try:
                    progress_window.destroy()
                except:
                    pass
            
            messagebox.showerror("Error", 
                               f"Failed to create PDF:\n{e}\n\n"
                               f"Check LOGS/ folder for details.",
                               parent=self.window)
    
    def _excel_to_pdf(self, excel_files: List[Path], pdf_path: Path, progress_label: tk.Label):
        """Convert Excel files to PDF - preserves exact Excel formatting
        
        Returns:
            List[Path]: List of PDF file paths created (one per excel file)
        """
        import tempfile
        
        # Try method 1: Excel COM automation (Windows + Microsoft Excel)
        if platform.system() == 'Windows':
            try:
                return self._excel_to_pdf_com(excel_files, pdf_path, progress_label)
            except Exception as e:
                print(f"Excel COM automation failed: {e}")
                print("Trying LibreOffice...")
        
        # Try method 2: LibreOffice UNO (cross-platform, works with LibreOffice)
        try:
            return self._excel_to_pdf_libreoffice(excel_files, pdf_path, progress_label)
        except Exception as e:
            print(f"LibreOffice conversion failed: {e}")
            print("Falling back to reportlab...")
        
        # Fallback: reportlab (cross-platform but basic formatting)
        return self._excel_to_pdf_reportlab(excel_files, pdf_path, progress_label)
    
    def _excel_to_pdf_com(self, excel_files: List[Path], pdf_path: Path, progress_label: tk.Label):
        """Convert Excel to PDF using Excel COM automation (Windows only)
        
        EXTREME optimization for 4GB RAM systems:
        - Process ONE file at a time (never load multiple)
        - Release Excel COM object after EACH file
        - Force garbage collection aggressively
        - Minimize temporary objects
        - Close workbooks immediately
        """
        import win32com.client
        import pythoncom
        import tempfile
        import gc
        from app.debug_logger import get_logger
        
        logger = get_logger()
        logger.section("Excel COM PDF Export (4GB RAM Optimized)")
        logger.info(f"Converting {len(excel_files)} files using Excel COM automation")
        logger.info("EXTREME memory optimization enabled for low-RAM systems")
        logger.start_timer("excel_com_total")
        
        # Initialize COM
        pythoncom.CoInitialize()
        logger.debug("COM initialized")
        
        saved_pdfs = []
        
        try:
            # EXTREME optimization: Start/stop Excel for EACH file
            # This ensures Excel doesn't accumulate memory
            for idx, excel_file in enumerate(excel_files, 1):
                logger.start_timer(f"convert_file_{idx}")
                logger.info(f"Processing file {idx}/{len(excel_files)}: {excel_file.name}")
                logger.log_memory_usage()
                
                progress_label.config(text=f"Converting {idx}/{len(excel_files)}: {excel_file.name}")
                progress_label.master.update()
                
                # Start Excel fresh for this file only
                excel = None
                wb = None
                
                try:
                    logger.start_timer(f"excel_startup_{idx}")
                    excel = win32com.client.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    excel.ScreenUpdating = False  # Performance: disable screen updates
                    excel.Calculation = -4135  # xlCalculationManual - disable auto-calculation
                    excel.EnableEvents = False  # Don't trigger events
                    logger.end_timer(f"excel_startup_{idx}")
                    logger.debug(f"Excel started for file {idx}")
                    
                    # Open workbook (read-only for performance)
                    logger.debug(f"Opening workbook: {excel_file}")
                    wb = excel.Workbooks.Open(str(excel_file.absolute()), ReadOnly=True, UpdateLinks=False, IgnoreReadOnlyRecommended=True)
                    
                    # Find PALLET SHEET
                    pallet_sheet = None
                    for sheet in wb.Sheets:
                        if sheet.Name.upper().replace(' ', '') == 'PALLETSHEET':
                            pallet_sheet = sheet
                            break
                    
                    if pallet_sheet:
                        logger.debug(f"Found PALLET SHEET: {pallet_sheet.Name}")
                        
                        # Export sheet to PDF
                        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                        temp_pdf.close()
                        
                        logger.start_timer(f"export_pdf_{idx}")
                        # Excel's ExportAsFixedFormat creates pixel-perfect PDF
                        pallet_sheet.ExportAsFixedFormat(
                            Type=0,  # xlTypePDF
                            Filename=temp_pdf.name,
                            Quality=0,  # xlQualityStandard (best quality)
                            IncludeDocProperties=False,  # Performance: skip properties
                            IgnorePrintAreas=False,
                            OpenAfterPublish=False  # Don't open PDF
                        )
                        logger.end_timer(f"export_pdf_{idx}")
                        
                        temp_pdf_path = temp_pdf.name
                        logger.debug(f"PDF created: {temp_pdf_path}")
                    else:
                        logger.warning(f"No PALLET SHEET found in {excel_file.name}")
                        temp_pdf_path = None
                    
                    # Close workbook immediately to free memory
                    if wb:
                        wb.Close(SaveChanges=False)
                        del wb  # Explicitly delete
                        logger.debug("Workbook closed and deleted")
                    
                    # CRITICAL: Quit Excel after EACH file for 4GB systems
                    if excel:
                        excel.Quit()
                        del excel  # Explicitly delete Excel COM object
                        logger.debug("Excel quit and deleted")
                    
                    # Move temp PDF to final location immediately (saves memory)
                    if temp_pdf_path:
                        import shutil
                        pdf_name = excel_file.stem + '.pdf'
                        final_pdf_path = pdf_path.parent / pdf_name
                        
                        # If file exists, add number suffix
                        counter = 1
                        while final_pdf_path.exists():
                            pdf_name = f"{excel_file.stem}_{counter}.pdf"
                            final_pdf_path = pdf_path.parent / pdf_name
                            counter += 1
                        
                        logger.debug(f"Moving {temp_pdf_path} -> {final_pdf_path}")
                        shutil.move(temp_pdf_path, final_pdf_path)
                        saved_pdfs.append(final_pdf_path)
                        del temp_pdf_path  # Free reference
                
                except Exception as e:
                    logger.error(f"Error converting {excel_file.name}: {e}", exc_info=e)
                    
                    # Clean up on error
                    if wb:
                        try:
                            wb.Close(SaveChanges=False)
                            del wb
                        except:
                            pass
                    if excel:
                        try:
                            excel.Quit()
                            del excel
                        except:
                            pass
                    continue
                    
                finally:
                    logger.end_timer(f"convert_file_{idx}")
                    
                    # AGGRESSIVE garbage collection for 4GB RAM
                    gc.collect()
                    gc.collect()  # Run twice for good measure
                    gc.collect()
                    
                    logger.log_memory_usage()
                    logger.info(f"File {idx}/{len(excel_files)} complete, memory released")
            
            if not saved_pdfs:
                logger.error("No PDFs were created!")
                raise Exception("No PDFs were created")
            
            logger.end_timer("excel_com_total")
            logger.info(f"Successfully created {len(saved_pdfs)} PDFs")
            logger.log_memory_usage()
            
            # Return list of individual PDFs (don't merge here)
            return saved_pdfs
                
        finally:
            pythoncom.CoUninitialize()
            logger.debug("COM uninitialized")
            # Final aggressive cleanup
            gc.collect()
            gc.collect()
            gc.collect()
    
    def _excel_to_pdf_libreoffice(self, excel_files: List[Path], pdf_path: Path, progress_label: tk.Label):
        """Convert Excel to PDF using LibreOffice (cross-platform)
        
        Optimized for low-end systems:
        - Processes files sequentially
        - Uses headless mode (no GUI overhead)
        - Minimal resource usage
        """
        import subprocess
        import tempfile
        import gc
        from app.debug_logger import get_logger
        
        logger = get_logger()
        logger.section("LibreOffice PDF Export")
        logger.info(f"Converting {len(excel_files)} files using LibreOffice")
        logger.start_timer("libreoffice_total")
        
        # Try to find LibreOffice
        libreoffice_paths = []
        
        if platform.system() == 'Windows':
            libreoffice_paths = [
                r"C:\Program Files\LibreOffice\program\soffice.exe",
                r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
                r"C:\Program Files\LibreOffice 7\program\soffice.exe",
                r"C:\Program Files\LibreOffice 24\program\soffice.exe",
            ]
        elif platform.system() == 'Darwin':  # macOS
            libreoffice_paths = [
                "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            ]
        else:  # Linux
            libreoffice_paths = [
                "/usr/bin/libreoffice",
                "/usr/bin/soffice",
            ]
        
        # Find available LibreOffice
        soffice = None
        for path in libreoffice_paths:
            if Path(path).exists():
                soffice = path
                break
        
        if not soffice:
            logger.error("LibreOffice not found on system")
            logger.debug(f"Searched paths: {libreoffice_paths}")
            raise Exception("LibreOffice not found on system")
        
        logger.info(f"Using LibreOffice: {soffice}")
        
        # Create temporary PDFs for each Excel file
        temp_pdfs = []
        temp_dir = tempfile.mkdtemp()
        
        try:
            for idx, excel_file in enumerate(excel_files, 1):
                logger.start_timer(f"convert_file_{idx}")
                logger.info(f"Processing file {idx}/{len(excel_files)}: {excel_file.name}")
                
                progress_label.config(text=f"Converting {idx}/{len(excel_files)}: {excel_file.name}")
                progress_label.master.update()
                
                # LibreOffice command to convert to PDF
                # --headless: run without GUI (performance)
                # --convert-to pdf: convert to PDF format
                # --outdir: output directory
                # --norestore: don't restore previous session (performance)
                cmd = [
                    soffice,
                    '--headless',
                    '--norestore',
                    '--convert-to', 'pdf',
                    '--outdir', temp_dir,
                    str(excel_file.absolute())
                ]
                
                logger.debug(f"Running command: {' '.join(cmd)}")
                
                # Run conversion
                try:
                    result = subprocess.run(
                        cmd,
                        capture_output=True,
                        text=True,
                        timeout=120  # 2 minute timeout per file (generous for low-end systems)
                    )
                    
                    if result.returncode != 0:
                        logger.error(f"LibreOffice conversion failed for {excel_file.name}")
                        logger.error(f"Return code: {result.returncode}")
                        logger.error(f"stderr: {result.stderr}")
                        logger.debug(f"stdout: {result.stdout}")
                        continue
                    
                    logger.debug(f"LibreOffice stdout: {result.stdout}")
                    
                except subprocess.TimeoutExpired:
                    logger.error(f"LibreOffice conversion timed out for {excel_file.name}")
                    continue
                
                # Find the generated PDF
                pdf_name = excel_file.stem + '.pdf'
                temp_pdf = Path(temp_dir) / pdf_name
                
                if temp_pdf.exists():
                    temp_pdfs.append(str(temp_pdf))
                    logger.debug(f"PDF created: {temp_pdf}")
                else:
                    logger.warning(f"PDF not generated for {excel_file.name}")
                    logger.debug(f"Expected location: {temp_pdf}")
                    logger.debug(f"Temp dir contents: {list(Path(temp_dir).glob('*'))}")
                
                logger.end_timer(f"convert_file_{idx}")
                # Force garbage collection after each file
                gc.collect()
                logger.log_memory_usage()
            
            if not temp_pdfs:
                logger.error("No PDFs were created by LibreOffice!")
                raise Exception("No PDFs were created by LibreOffice")
            
            # Save individual PDFs with proper names (one per pallet)
            import shutil
            saved_pdfs = []
            
            logger.info(f"Saving {len(temp_pdfs)} PDFs to final locations")
            for temp_pdf, excel_file in zip(temp_pdfs, excel_files):
                # Create filename based on Excel filename
                pdf_name = excel_file.stem + '.pdf'
                final_pdf_path = pdf_path.parent / pdf_name
                
                # If file exists, add number suffix
                counter = 1
                while final_pdf_path.exists():
                    pdf_name = f"{excel_file.stem}_{counter}.pdf"
                    final_pdf_path = pdf_path.parent / pdf_name
                    counter += 1
                
                # Copy temp PDF to final location
                logger.debug(f"Copying {temp_pdf} -> {final_pdf_path}")
                shutil.copy(temp_pdf, final_pdf_path)
                saved_pdfs.append(final_pdf_path)
            
            logger.end_timer("libreoffice_total")
            logger.info(f"Successfully created {len(saved_pdfs)} PDFs")
            
            # Return list of individual PDFs (don't merge here)
            return saved_pdfs
                
        finally:
            # Clean up temp directory
            import shutil
            logger.debug(f"Cleaning up temp directory: {temp_dir}")
            try:
                shutil.rmtree(temp_dir)
                logger.debug("Temp directory deleted")
            except Exception as e:
                logger.warning(f"Failed to delete temp directory: {e}")
            
            gc.collect()  # Final cleanup
    
    def _excel_to_pdf_reportlab(self, excel_files: List[Path], pdf_path: Path, progress_label: tk.Label):
        """Convert Excel files to PDF using reportlab (fallback method)"""
        from openpyxl import load_workbook
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.pdfgen import canvas
        
        c = canvas.Canvas(str(pdf_path), pagesize=landscape(letter))
        page_width, page_height = landscape(letter)
        
        for idx, excel_file in enumerate(excel_files, 1):
            progress_label.config(text=f"Processing {idx}/{len(excel_files)}: {excel_file.name}")
            progress_label.master.update()
            
            try:
                # Load workbook
                wb = load_workbook(excel_file, read_only=True, data_only=True)
                
                # Get PALLET SHEET
                sheet_name = None
                for name in wb.sheetnames:
                    if name.upper().replace(' ', '') == 'PALLETSHEET':
                        sheet_name = name
                        break
                
                if not sheet_name:
                    wb.close()
                    continue  # Skip if no PALLET SHEET
                
                ws = wb[sheet_name]
                
                # Add new page for each Excel file
                if idx > 1:
                    c.showPage()
                
                # Add title
                c.setFont("Helvetica-Bold", 16)
                c.drawString(50, page_height - 50, f"Pallet Sheet: {excel_file.stem}")
                
                # Read and display key information
                c.setFont("Helvetica", 10)
                y_position = page_height - 100
                
                # Key cells
                key_info = []
                try:
                    b1 = ws['B1'].value
                    if b1:
                        key_info.append(f"Panel Type: {b1}")
                except:
                    pass
                
                try:
                    b3 = ws['B3'].value
                    if b3:
                        key_info.append(f"Pallet ID: {b3}")
                except:
                    pass
                
                try:
                    g3 = ws['G3'].value
                    if g3:
                        key_info.append(f"Date: {g3}")
                except:
                    pass
                
                for info in key_info:
                    c.drawString(50, y_position, info)
                    y_position -= 20
                
                # List serial numbers
                y_position -= 20
                c.setFont("Helvetica-Bold", 11)
                c.drawString(50, y_position, "Serial Numbers:")
                y_position -= 25
                c.setFont("Helvetica", 9)
                
                serials_found = 0
                for row in range(5, 31):  # Rows 5-30
                    try:
                        serial = ws[f'B{row}'].value
                        if serial:
                            c.drawString(70, y_position, f"{serials_found + 1:2d}. {serial}")
                            y_position -= 18
                            serials_found += 1
                            if y_position < 100:
                                c.showPage()
                                y_position = page_height - 50
                                c.setFont("Helvetica", 9)
                    except:
                        continue
                
                wb.close()
                
            except Exception as e:
                # Continue with next file if one fails
                continue
        
        c.save()
        
        # For reportlab, we created one merged PDF already
        # Return it as a single-item list for consistency
        return [pdf_path]
    
    def _merge_pdfs(self, pdf_files: List[str], output_path: Path):
        """Merge multiple PDF files into one"""
        try:
            from PyPDF2 import PdfMerger
            
            merger = PdfMerger()
            for pdf_file in pdf_files:
                merger.append(pdf_file)
            
            merger.write(str(output_path))
            merger.close()
            
        except ImportError:
            # PyPDF2 not available, try pypdf instead
            try:
                from pypdf import PdfMerger
                
                merger = PdfMerger()
                for pdf_file in pdf_files:
                    merger.append(pdf_file)
                
                merger.write(str(output_path))
                merger.close()
                
            except ImportError:
                # No PDF merger available, just use the first PDF
                import shutil
                if pdf_files:
                    shutil.copy(pdf_files[0], output_path)
    
    def _print_pdf(self, pdf_path: Path):
        """Print PDF file - automatically opens print dialog for the saved PDF"""
        try:
            system = platform.system()
            if system == 'Windows':
                # Windows: Try multiple approaches to open print dialog
                print_dialog_opened = False
                
                # Method 1: Check for bundled SumatraPDF first (PRIORITY)
                # This is bundled with the app for reliable print dialog
                bundled_sumatra = None
                try:
                    # Check if running from PyInstaller bundle
                    if getattr(sys, 'frozen', False):
                        print(f"DEBUG: Running as frozen exe")
                        # Running as compiled exe
                        if hasattr(sys, '_MEIPASS'):
                            # PyInstaller onefile - extract from temp
                            bundled_sumatra = Path(sys._MEIPASS) / 'external_tools' / 'SumatraPDF.exe'
                            print(f"DEBUG: Onefile mode - checking: {bundled_sumatra}")
                        else:
                            # PyInstaller onedir
                            bundled_sumatra = Path(sys.executable).parent / 'external_tools' / 'SumatraPDF.exe'
                            print(f"DEBUG: Onedir mode - checking: {bundled_sumatra}")
                    else:
                        print(f"DEBUG: Running from source (not frozen)")
                    
                    if bundled_sumatra:
                        print(f"DEBUG: Bundled SumatraPDF path: {bundled_sumatra}")
                        print(f"DEBUG: Exists? {bundled_sumatra.exists()}")
                        if bundled_sumatra.exists():
                            print(f"DEBUG: Attempting to launch SumatraPDF...")
                            try:
                                subprocess.Popen([str(bundled_sumatra), '-print-dialog', str(pdf_path.absolute())],
                                               creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
                                print_dialog_opened = True
                                print(f"DEBUG: SumatraPDF launched successfully!")
                                messagebox.showinfo(
                                    "PDF Created & Print Dialog Opening",
                                    f"PDF saved to:\n{pdf_path}\n\n"
                                    f"Print dialog is opening...\n"
                                    f"Adjust settings and click Print.",
                                    parent=self.window
                                )
                            except Exception as e:
                                print(f"DEBUG: Bundled SumatraPDF failed to launch: {e}")
                        else:
                            print(f"DEBUG: Bundled SumatraPDF not found at expected location")
                except Exception as e:
                    print(f"DEBUG: Error checking for bundled SumatraPDF: {e}")
                
                # Method 2: Check for installed PDF viewers with print flags
                if not print_dialog_opened:
                    pdf_viewers = [
                        # Installed SumatraPDF
                        (r"C:\Program Files\SumatraPDF\SumatraPDF.exe", ['-print-dialog']),
                        (r"C:\Program Files (x86)\SumatraPDF\SumatraPDF.exe", ['-print-dialog']),
                        # Adobe Reader with /p (print) flag
                        (r"C:\Program Files\Adobe\Acrobat DC\Acrobat\Acrobat.exe", ['/p']),
                        (r"C:\Program Files (x86)\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe", ['/p']),
                        (r"C:\Program Files\Adobe\Acrobat Reader DC\Reader\AcroRd32.exe", ['/p']),
                    ]
                    
                    for viewer_path, flags in pdf_viewers:
                        if Path(viewer_path).exists():
                            try:
                                subprocess.Popen([viewer_path] + flags + [str(pdf_path.absolute())],
                                               creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
                                print_dialog_opened = True
                                messagebox.showinfo(
                                    "PDF Created & Print Dialog Opening",
                                    f"PDF saved to:\n{pdf_path}\n\n"
                                    f"Print dialog is opening...\n"
                                    f"Adjust settings and click Print.",
                                    parent=self.window
                                )
                                break
                            except Exception:
                                continue
                
                # Method 2: Use Windows Edge browser with print flag (always available on Win10/11)
                if not print_dialog_opened:
                    try:
                        # Microsoft Edge can open PDFs and has a --print flag
                        edge_path = r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe"
                        if not Path(edge_path).exists():
                            edge_path = r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"
                        
                        if Path(edge_path).exists():
                            # Open PDF in Edge with print dialog
                            subprocess.Popen([edge_path, '--print', str(pdf_path.absolute())],
                                           creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
                            print_dialog_opened = True
                        messagebox.showinfo(
                                "PDF Created & Print Dialog Opening",
                                f"PDF saved to:\n{pdf_path}\n\n"
                                f"Print dialog is opening in Edge...",
                            parent=self.window
                        )
                    except Exception:
                        pass
                
                # Fallback: Just open PDF and guide user
                if not print_dialog_opened:
                    try:
                        os.startfile(str(pdf_path.absolute()))
                        messagebox.showinfo(
                            "PDF Created - Manual Print",
                            f"PDF saved to:\n{pdf_path}\n\n"
                            f"✓ PDF opened in your default viewer\n\n"
                            f"To print:\n"
                            f"1. Press Ctrl+P in the PDF viewer\n"
                            f"2. Adjust print settings\n"
                            f"3. Click Print\n\n"
                            f"(Automatic print dialog requires Adobe Reader,\n"
                            f"SumatraPDF, or Microsoft Edge)",
                            parent=self.window
                        )
                    except Exception:
                        messagebox.showwarning(
                            "PDF Saved",
                            f"PDF saved to:\n{pdf_path}\n\n"
                            f"Please navigate to the file and right-click → Print",
                            parent=self.window
                        )
            elif system == 'Darwin':  # macOS
                # macOS: Open PDF and trigger print dialog with AppleScript
                try:
                    # Open in Preview
                    subprocess.Popen(['open', '-a', 'Preview', str(pdf_path.absolute())])

                    # Wait for Preview to open, then trigger print dialog
                    import time
                    time.sleep(0.5)
                    
                    applescript = f'''
                        tell application "Preview"
                            activate
                            delay 0.3
                        end tell
                        tell application "System Events"
                            tell process "Preview"
                                keystroke "p" using command down
                            end tell
                        end tell
                        '''
                    subprocess.Popen(['osascript', '-e', applescript])

                    messagebox.showinfo(
                        "PDF Created & Print Dialog Opened",
                        f"PDF saved to:\n{pdf_path}\n\n"
                        f"Print dialog should now be open in Preview.",
                        parent=self.window
                    )
                except Exception as e:
                    # Fallback: Just open in Preview
                    subprocess.run(['open', '-a', 'Preview', str(pdf_path.absolute())], check=False)
                    messagebox.showinfo(
                        "PDF Created",
                        f"PDF saved to:\n{pdf_path}\n\n"
                        f"Opened in Preview. Use Cmd+P to print.",
                            parent=self.window
                        )
            else:  # Linux
                # Linux: Use lp command to send to printer with dialog
                try:
                    # Try to use lpr which shows print dialog on some systems
                    subprocess.Popen(['lpr', str(pdf_path.absolute())])
                    messagebox.showinfo(
                        "PDF Created & Sent to Printer",
                        f"PDF saved to:\n{pdf_path}\n\n"
                        f"Sent to default printer.",
                        parent=self.window
                    )
                except Exception:
                    # Fallback: Open with default viewer
                    subprocess.run(['xdg-open', str(pdf_path.absolute())], check=False)
                    messagebox.showinfo(
                        "PDF Created",
                        f"PDF saved to:\n{pdf_path}\n\n"
                        f"Opened in default viewer. Use Ctrl+P to print.",
                        parent=self.window
                    )
        except Exception as e:
            messagebox.showerror("Error", f"Could not open print dialog:\n{e}", 
                               parent=self.window)
    
    def _print_excel_files(self, excel_files: List[Path]):
        """Print Excel files directly (fallback if reportlab not available)"""
        try:
            system = platform.system()
            for excel_file in excel_files:
                if system == 'Windows':
                    subprocess.run(['start', '/MIN', str(excel_file.absolute())], 
                                 shell=True, check=False)
                elif system == 'Darwin':  # macOS
                    subprocess.run(['open', str(excel_file.absolute())], check=False)
                else:  # Linux
                    subprocess.run(['xdg-open', str(excel_file.absolute())], check=False)
            
            messagebox.showinfo(
                "Print",
                f"Opened {len(excel_files)} file(s) for printing.\n\n"
                "Please use File > Print in Excel/Preview to print.",
                parent=self.window
            )
        except Exception as e:
            messagebox.showerror("Error", f"Could not open files for printing:\n{e}", 
                               parent=self.window)
    
    def _update_customer_filter_options(self):
        """Update customer filter dropdown with available customers"""
        if not self.customer_manager:
            return
        
        # Get current selection
        current_selection = self.customer_filter_var.get()
        
        # Get all customers
        self.customer_manager.refresh_customers()
        customer_names = self.customer_manager.get_customer_names()
        
        # Build menu options: "ALL" + customer names
        menu = self.customer_filter_menu['menu']
        menu.delete(0, tk.END)
        
        # Add "ALL" option
        menu.add_command(label="ALL", command=lambda: self._set_customer_filter("ALL"))
        
        # Add each customer
        for customer_name in customer_names:
            menu.add_command(label=customer_name, command=lambda name=customer_name: self._set_customer_filter(name))
        
        # Restore selection if it still exists
        if current_selection in ["ALL"] + customer_names:
            self.customer_filter_var.set(current_selection)
        else:
            self.customer_filter_var.set("ALL")
    
    def _set_customer_filter(self, value):
        """Set customer filter value"""
        self.customer_filter_var.set(value)
        self.load_history()
    
    def _on_customer_filter_changed(self, *args):
        """Handle customer filter change"""
        # Only load if tree widget exists (avoid error during initialization)
        if hasattr(self, 'tree') and self.tree:
            self.load_history()
    
    def _on_search_changed(self, *args):
        """Handle barcode search change (debounced)"""
        # Debounce search to avoid too many updates while typing
        if hasattr(self, '_search_timer'):
            self.window.after_cancel(self._search_timer)
        
        self._search_timer = self.window.after(300, self.load_history)  # Wait 300ms after typing stops
    
    def _sort_by_column(self, column):
        """Sort table by the specified column"""
        # Check if tree has any items
        tree_items = list(self.tree.get_children())
        if not tree_items:
            return  # Nothing to sort
        
        # Toggle sort direction if clicking the same column
        if self.sort_column == column:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = column
            self.sort_reverse = False
        
        # Get all items from the tree
        items = [(self.tree.item(item, 'values'), item) for item in tree_items]
        
        # Determine sort key based on column
        if column == "Pallet #":
            # Sort by pallet number (column index 1)
            sort_key = lambda x: int(x[0][1]) if x[0][1] and str(x[0][1]).isdigit() else 0
        elif column == "Completed":
            # Sort by completed date (column index 2)
            def parse_date(value):
                if not value or value == 'N/A':
                    return datetime.min
                try:
                    # Try parsing the date part (format: "YYYY-MM-DD HH:MM:SS")
                    date_str = str(value).split()[0] if ' ' in str(value) else str(value)
                    return datetime.strptime(date_str, "%Y-%m-%d")
                except:
                    return datetime.min
            sort_key = lambda x: parse_date(x[0][2])
        elif column == "File Name":
            # Sort by file name (column index 3)
            sort_key = lambda x: str(x[0][3]) if len(x[0]) > 3 else ""
        else:
            return  # Unknown column
        
        # Sort items
        items.sort(key=sort_key, reverse=self.sort_reverse)
        
        # Reorder items in tree
        for values, item in items:
            self.tree.move(item, "", "end")
        
        # Update header to show sort indicator
        self._update_sort_indicators()
    
    def _update_sort_indicators(self):
        """Update column headers to show sort direction"""
        columns = ["Pallet #", "Completed", "File Name"]
        for col in columns:
            current_text = self.tree.heading(col, "text")
            # Remove existing sort indicators
            base_text = current_text.replace(" ▲", "").replace(" ▼", "")
            
            if self.sort_column == col:
                indicator = " ▼" if self.sort_reverse else " ▲"
                self.tree.heading(col, text=base_text + indicator)
            else:
                self.tree.heading(col, text=base_text)
    
    def _auto_select_single_pallet(self, pallet):
        """Automatically select a pallet when search narrows down to one result"""
        try:
            # Find the tree item for this pallet
            for item in self.tree.get_children():
                if self.item_to_pallet.get(item) is pallet:
                    # Select this item (Windows File Explorer style)
                    self.tree.selection_set(item)
                    self.tree.focus(item)
                    self.tree.see(item)  # Scroll into view
                    
                    # Update checkbox state
                    self.checkbox_states[item] = True
                    self._update_checkbox_display(item)
                    
                    # Show details
                    self._show_pallet_details(item)
                    break
        except Exception as e:
            print(f"DEBUG: Error auto-selecting pallet: {e}")
